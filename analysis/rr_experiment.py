from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

if __package__ is None or __package__ == "":
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from analysis.backtest import run_backtest
from analysis.candidate_generator import generate_candidates
from analysis.diagnostics import DEFAULT_THRESHOLDS, threshold_diagnostics
from analysis.economic_calendar import EconomicEvent, load_economic_calendar, parse_currencies
from analysis.market_data import MarketHistory, load_history
from analysis.models import BacktestResult, Candidate
from analysis.reports import candidate_result_rows, summarize


VARIABLE_POLICIES = ("setup_ladder", "space_ladder", "side_ladder")
DEFAULT_RR_VALUES = (2.0, 3.0, 4.0, 5.0)
DEFAULT_MAX_SIDE_POSITIVE_SHARE = 0.85
DEFAULT_MIN_ADOPTION_PF = 1.0


def run_rr_experiment(
    history: MarketHistory,
    *,
    rr_values: list[float],
    side: str = "both",
    min_score: float | None = None,
    max_hold_minutes: int = 60,
    swing_left: int = 3,
    swing_right: int = 3,
    min_atr_distance: float = 0.5,
    max_risk_atr: float = 3.0,
    score_profile: str = "side",
    exclude_blackout_times: bool = True,
    blackout_events: list[EconomicEvent] | tuple[EconomicEvent, ...] = (),
    news_before_minutes: int = 10,
    news_after_minutes: int = 10,
    news_min_impact: str = "high",
    news_currencies: tuple[str, ...] | list[str] | None = ("USD", "XAU", "ALL", "*"),
    variable_policies: list[str] | None = None,
) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    rr_values = sorted(set(rr_values))
    spread_price = history.spread_points * history.point
    candidate_sets = {
        rr: generate_candidates(
            history,
            risk_reward=rr,
            swing_left=swing_left,
            swing_right=swing_right,
            min_atr_distance=min_atr_distance,
            max_risk_atr=max_risk_atr,
            min_score=None,
            score_profile=score_profile,
            exclude_blackout_times=exclude_blackout_times,
            blackout_events=blackout_events,
            news_before_minutes=news_before_minutes,
            news_after_minutes=news_after_minutes,
            news_min_impact=news_min_impact,
            news_currencies=news_currencies,
        )
        for rr in rr_values
    }

    summary_rows: list[dict[str, object]] = []
    threshold_rows: list[dict[str, object]] = []
    distribution_rows: list[dict[str, object]] = []
    detail_rows: list[dict[str, object]] = []

    for rr, candidates in candidate_sets.items():
        side_candidates = _filter_side(candidates, side)
        threshold_results = run_backtest(side_candidates, history.bars("M1"), max_hold_minutes=max_hold_minutes, spread_price=spread_price)
        for row in threshold_diagnostics(side_candidates, threshold_results, thresholds=DEFAULT_THRESHOLDS):
            threshold_rows.append({"strategy": f"fixed_1_{rr:g}", "policy": "fixed", "rr": rr, **row})

        selected = _filter_min_score(side_candidates, min_score)
        results = run_backtest(selected, history.bars("M1"), max_hold_minutes=max_hold_minutes, spread_price=spread_price)
        summary_rows.append(_summary_row(f"fixed_1_{rr:g}", "fixed", rr, side, min_score, selected, results))
        detail_rows.extend(_detail_rows(f"fixed_1_{rr:g}", selected, results))

    for policy in variable_policies or []:
        selected_all = select_variable_rr_candidates(candidate_sets, policy)
        selected = _filter_min_score(_filter_side(selected_all, side), min_score)
        results = run_backtest(selected, history.bars("M1"), max_hold_minutes=max_hold_minutes, spread_price=spread_price)
        summary_rows.append(_summary_row(f"variable_{policy}", policy, "variable", side, min_score, selected, results))
        threshold_results = run_backtest(_filter_side(selected_all, side), history.bars("M1"), max_hold_minutes=max_hold_minutes, spread_price=spread_price)
        for row in threshold_diagnostics(_filter_side(selected_all, side), threshold_results, thresholds=DEFAULT_THRESHOLDS):
            threshold_rows.append({"strategy": f"variable_{policy}", "policy": policy, "rr": "variable", **row})
        distribution_rows.extend(_rr_distribution_rows(f"variable_{policy}", policy, selected))
        detail_rows.extend(_detail_rows(f"variable_{policy}", selected, results))

    summary_rows.sort(key=lambda row: (float(row.get("avg_r", 0.0) or 0.0), float(row.get("pf", 0.0) or 0.0)), reverse=True)
    threshold_rows.sort(key=lambda row: (str(row.get("strategy", "")), float(row.get("threshold", 0.0) or 0.0)))
    return summary_rows, threshold_rows, distribution_rows, detail_rows


def rr_strategy_adoption_audit(
    summary_rows: list[dict[str, object]],
    threshold_rows: list[dict[str, object]],
    detail_rows: list[dict[str, object]],
    *,
    min_pf: float = DEFAULT_MIN_ADOPTION_PF,
    max_side_positive_share: float = DEFAULT_MAX_SIDE_POSITIVE_SHARE,
    monotonic_tolerance: float = 0.0,
) -> list[dict[str, object]]:
    fixed_rows = [row for row in summary_rows if row.get("policy") == "fixed"]
    best_fixed = max(fixed_rows, key=_balance_score) if fixed_rows else None
    rows: list[dict[str, object]] = []
    for summary in sorted(summary_rows, key=lambda row: str(row.get("strategy", ""))):
        strategy = str(summary.get("strategy") or "")
        policy = str(summary.get("policy") or "")
        side = str(summary.get("side") or "both")
        strategy_thresholds = [row for row in threshold_rows if row.get("strategy") == strategy]
        strategy_details = [row for row in detail_rows if row.get("strategy") == strategy]
        score_audit = score_threshold_audit(strategy_thresholds, tolerance=monotonic_tolerance)
        side_audit = side_balance_audit(
            strategy_details,
            strategy_side=side,
            max_side_positive_share=max_side_positive_share,
        )

        reasons: list[str] = []
        count = _to_int(summary.get("count"))
        avg_r = _to_float(summary.get("avg_r"))
        pf = _to_float(summary.get("pf"))
        if count <= 0:
            reasons.append("no_trades")
        if avg_r <= 0.0:
            reasons.append("avg_r_not_positive")
        if pf < min_pf:
            reasons.append(f"pf_below_{min_pf:g}")
        if not score_audit["ok"]:
            reasons.append(str(score_audit["status"]))
        if not side_audit["ok"]:
            reasons.extend(str(reason) for reason in side_audit["reasons"])

        best_fixed_strategy = ""
        best_fixed_balance_score: float | str = ""
        fixed_balance_delta: float | str = ""
        if policy != "fixed":
            if best_fixed is None:
                reasons.append("missing_fixed_rr_baseline")
            else:
                best_fixed_strategy = str(best_fixed.get("strategy") or "")
                current_score = _balance_score(summary)
                best_score = _balance_score(best_fixed)
                best_fixed_balance_score = round(best_score, 6)
                fixed_balance_delta = round(current_score - best_score, 6)
                if current_score + monotonic_tolerance < best_score:
                    reasons.append("variable_balance_below_best_fixed")

        rows.append(
            {
                "strategy": strategy,
                "policy": policy,
                "side": side,
                "status": "candidate" if not reasons else "rejected",
                "reasons": reasons,
                "count": count,
                "avg_r": avg_r,
                "pf": pf,
                "total_r": _to_float(summary.get("total_r")),
                "max_drawdown_r": _to_float(summary.get("max_drawdown_r")),
                "balance_score": round(_balance_score(summary), 6),
                "best_fixed_strategy": best_fixed_strategy,
                "best_fixed_balance_score": best_fixed_balance_score,
                "fixed_balance_delta": fixed_balance_delta,
                "score_threshold_status": score_audit["status"],
                "score_threshold_degradation_count": score_audit["degradation_count"],
                "score_threshold_degradations": score_audit["degradations"],
                "side_balance_status": side_audit["status"],
                "side_positive_share": side_audit["max_positive_share"],
                "side_positive_share_side": side_audit["max_positive_share_side"],
                "side_rows": side_audit["side_rows"],
            }
        )
    return rows


def score_threshold_audit(
    threshold_rows: list[dict[str, object]],
    *,
    tolerance: float = 0.0,
) -> dict[str, object]:
    rows = sorted(
        [row for row in threshold_rows if _to_int(row.get("count")) > 0],
        key=lambda row: _to_float(row.get("threshold")),
    )
    degradations: list[dict[str, object]] = []
    previous: dict[str, object] | None = None
    for row in rows:
        if previous is not None:
            previous_avg_r = _to_float(previous.get("avg_r"))
            previous_pf = _to_float(previous.get("pf"))
            current_avg_r = _to_float(row.get("avg_r"))
            current_pf = _to_float(row.get("pf"))
            fields: list[str] = []
            if current_avg_r + tolerance < previous_avg_r:
                fields.append("avg_r")
            if current_pf + tolerance < previous_pf:
                fields.append("pf")
            if fields:
                degradations.append(
                    {
                        "from_threshold": previous.get("threshold"),
                        "to_threshold": row.get("threshold"),
                        "fields": fields,
                        "from_avg_r": previous_avg_r,
                        "to_avg_r": current_avg_r,
                        "from_pf": previous_pf,
                        "to_pf": current_pf,
                    }
                )
        previous = row
    if len(rows) < 2:
        status = "insufficient_score_threshold_rows"
    elif degradations:
        status = "score_threshold_degrades"
    else:
        status = "score_threshold_non_degrading"
    return {
        "ok": status == "score_threshold_non_degrading",
        "status": status,
        "checked_threshold_count": len(rows),
        "degradation_count": len(degradations),
        "degradations": degradations,
    }


def side_balance_audit(
    detail_rows: list[dict[str, object]],
    *,
    strategy_side: str,
    max_side_positive_share: float = DEFAULT_MAX_SIDE_POSITIVE_SHARE,
) -> dict[str, object]:
    side_rows = [_side_summary(side, rows) for side, rows in sorted(_group_detail_rows(detail_rows, "side").items())]
    if strategy_side != "both":
        return {
            "ok": True,
            "status": "single_side_not_applicable",
            "reasons": [],
            "max_positive_share": "",
            "max_positive_share_side": "",
            "side_rows": side_rows,
        }

    reasons: list[str] = []
    by_side = {str(row["side"]): row for row in side_rows}
    for required_side in ("buy", "sell"):
        if required_side not in by_side:
            reasons.append(f"{required_side}_side_missing")
            continue
        if _to_float(by_side[required_side].get("avg_r")) < 0.0:
            reasons.append(f"{required_side}_avg_r_negative")

    positive_totals = {
        str(row["side"]): max(_to_float(row.get("total_r")), 0.0)
        for row in side_rows
    }
    total_positive = sum(positive_totals.values())
    if total_positive > 0:
        max_side, max_total = max(positive_totals.items(), key=lambda item: item[1])
        max_share: float | str = round(max_total / total_positive, 4)
        if max_share > max_side_positive_share:
            reasons.append(f"positive_total_r_side_share_gt_{max_side_positive_share:g}")
    else:
        max_side = ""
        max_share = ""

    return {
        "ok": not reasons,
        "status": "side_balance_ok" if not reasons else "side_balance_failed",
        "reasons": reasons,
        "max_positive_share": max_share,
        "max_positive_share_side": max_side,
        "side_rows": side_rows,
    }


def write_experiment_json(
    path: str | Path,
    summary_rows: list[dict[str, object]],
    threshold_rows: list[dict[str, object]],
    distribution_rows: list[dict[str, object]],
    detail_rows: list[dict[str, object]],
    audit_rows: list[dict[str, object]],
    *,
    metadata: dict[str, object] | None = None,
) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    run_metadata = metadata or {}
    payload = {
        "ok": True,
        "generated_at": datetime.now().strftime("%Y.%m.%d %H:%M:%S"),
        "metadata": run_metadata,
        "summary_rows": summary_rows,
        "threshold_rows": threshold_rows,
        "distribution_rows": distribution_rows,
        "detail_rows": detail_rows,
        "adoption_audit": audit_rows,
    }
    payload.update(_metadata_top_level_aliases(run_metadata))
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_experiment_markdown(
    path: str | Path,
    summary_rows: list[dict[str, object]],
    threshold_rows: list[dict[str, object]],
    distribution_rows: list[dict[str, object]],
    audit_rows: list[dict[str, object]],
    *,
    metadata: dict[str, object] | None = None,
) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        format_experiment_markdown(summary_rows, threshold_rows, distribution_rows, audit_rows, metadata=metadata),
        encoding="utf-8",
    )


def format_experiment_markdown(
    summary_rows: list[dict[str, object]],
    threshold_rows: list[dict[str, object]],
    distribution_rows: list[dict[str, object]],
    audit_rows: list[dict[str, object]],
    *,
    metadata: dict[str, object] | None = None,
) -> str:
    lines = [
        "# RR Strategy Experiment",
        "",
        "## Run Metadata",
        "",
    ]
    _append_markdown_table(lines, [_metadata_markdown_row(metadata or {})] if metadata else [])
    lines.extend([
        "",
        "## Adoption Audit",
        "",
    ])
    _append_markdown_table(lines, audit_rows)
    lines.extend(["", "## Strategy Summary", ""])
    _append_markdown_table(lines, summary_rows)
    lines.extend(["", "## Variable RR Distribution", ""])
    _append_markdown_table(lines, distribution_rows)
    lines.extend(["", "## Threshold Diagnostics", ""])
    _append_markdown_table(lines, threshold_rows)
    return "\n".join(lines) + "\n"


def experiment_metadata(
    args: argparse.Namespace,
    history: MarketHistory,
    calendar_events: list[EconomicEvent] | tuple[EconomicEvent, ...],
) -> dict[str, object]:
    return {
        "history": str(args.history),
        "history_symbol": history.symbol,
        "history_server_time": history.server_time,
        "history_hours": history.history_hours,
        "history_timeframes": _timeframe_metadata(history),
        "point": history.point,
        "spread_points": history.spread_points,
        "rr_values": list(args.rr_values),
        "variable_policies": list(args.variable_policies),
        "side": args.side,
        "min_score": args.min_score,
        "max_hold_minutes": args.max_hold_minutes,
        "swing_left": args.swing_left,
        "swing_right": args.swing_right,
        "min_atr_distance": args.min_atr_distance,
        "max_risk_atr": args.max_risk_atr,
        "score_profile": args.score_profile,
        "include_blackout_times": args.include_blackout_times,
        "exclude_blackout_times": not args.include_blackout_times,
        "calendar": str(args.calendar),
        "calendar_event_count": len(calendar_events),
        "calendar_input_utc_offset": args.calendar_input_utc_offset,
        "calendar_server_utc_offset": args.calendar_server_utc_offset,
        "news_before_minutes": args.news_before_minutes,
        "news_after_minutes": args.news_after_minutes,
        "news_min_impact": args.news_min_impact,
        "news_currencies": list(parse_currencies(args.news_currencies)),
        "adoption_min_pf": DEFAULT_MIN_ADOPTION_PF,
        "adoption_max_side_positive_share": DEFAULT_MAX_SIDE_POSITIVE_SHARE,
        "fill_model": "tp_sl_first_touch_same_bar_sl_first",
    }


def _timeframe_metadata(history: MarketHistory) -> dict[str, dict[str, object]]:
    rows: dict[str, dict[str, object]] = {}
    for timeframe, bars in sorted(history.timeframes.items()):
        rows[timeframe] = {
            "bars": len(bars),
            "first_time": bars[0].time_text if bars else "",
            "last_time": bars[-1].time_text if bars else "",
        }
    return rows


def _metadata_top_level_aliases(metadata: dict[str, object]) -> dict[str, object]:
    keys = (
        "history",
        "history_symbol",
        "history_server_time",
        "history_hours",
        "rr_values",
        "variable_policies",
        "side",
        "min_score",
        "max_hold_minutes",
        "score_profile",
        "exclude_blackout_times",
        "calendar",
        "calendar_event_count",
    )
    return {key: metadata[key] for key in keys if key in metadata}


def _metadata_markdown_row(metadata: dict[str, object]) -> dict[str, object]:
    timeframes = metadata.get("history_timeframes")
    m1 = timeframes.get("M1", {}) if isinstance(timeframes, dict) else {}
    return {
        "history": metadata.get("history", ""),
        "symbol": metadata.get("history_symbol", ""),
        "server_time": metadata.get("history_server_time", ""),
        "history_hours": metadata.get("history_hours", ""),
        "M1_bars": m1.get("bars", "") if isinstance(m1, dict) else "",
        "M1_first": m1.get("first_time", "") if isinstance(m1, dict) else "",
        "M1_last": m1.get("last_time", "") if isinstance(m1, dict) else "",
        "rr_values": metadata.get("rr_values", ""),
        "variable_policies": metadata.get("variable_policies", ""),
        "side": metadata.get("side", ""),
        "min_score": metadata.get("min_score", ""),
        "max_hold_minutes": metadata.get("max_hold_minutes", ""),
        "score_profile": metadata.get("score_profile", ""),
        "exclude_blackout_times": metadata.get("exclude_blackout_times", ""),
        "calendar_events": metadata.get("calendar_event_count", ""),
    }


def select_variable_rr_candidates(candidate_sets: dict[float, list[Candidate]], policy: str) -> list[Candidate]:
    if policy not in VARIABLE_POLICIES:
        raise ValueError(f"unknown variable RR policy: {policy}")

    rr_values = sorted(candidate_sets)
    by_id: dict[str, dict[float, Candidate]] = {}
    for rr, candidates in candidate_sets.items():
        for candidate in candidates:
            by_id.setdefault(candidate.candidate_id, {})[rr] = candidate

    selected: list[Candidate] = []
    for candidate_id in sorted(by_id):
        choices = by_id[candidate_id]
        if policy == "setup_ladder":
            chosen = _choose_setup_ladder(choices, rr_values)
        elif policy == "space_ladder":
            chosen = _choose_space_ladder(choices, rr_values)
        else:
            chosen = _choose_side_ladder(choices, rr_values)
        if chosen is not None:
            selected.append(chosen)
    return selected


def setup_quality_score(candidate: Candidate) -> float:
    parts = candidate.score_parts
    rr_component = min(candidate.risk_reward / 5.0, 1.0) * _rr_component_weight(candidate)
    risk_without_rr = max(_num(parts.get("risk_reward_score")) - rr_component, 0.0)
    return (
        _num(parts.get("trend_score"))
        + _num(parts.get("structure_score"))
        + _num(parts.get("entry_trigger_score"))
        + risk_without_rr
        + _num(parts.get("cost_penalty"))
        + _num(parts.get("chop_penalty"))
    )


def _rr_component_weight(candidate: Candidate) -> float:
    profile = str(candidate.features.get("score_profile") or "side")
    if profile in ("side", "buy", "sell"):
        return 8.0
    return 12.0


def parse_rr_values(value: str) -> list[float]:
    values: list[float] = []
    for item in value.split(","):
        item = item.strip()
        if not item:
            continue
        parsed = float(item)
        if parsed <= 0:
            raise argparse.ArgumentTypeError("RR values must be positive")
        values.append(parsed)
    if not values:
        raise argparse.ArgumentTypeError("at least one RR value is required")
    return values


def parse_policies(value: str) -> list[str]:
    if value.lower() in ("", "none", "off"):
        return []
    policies = [item.strip() for item in value.split(",") if item.strip()]
    unknown = [policy for policy in policies if policy not in VARIABLE_POLICIES]
    if unknown:
        raise argparse.ArgumentTypeError(f"unknown variable policies: {', '.join(unknown)}")
    return policies


def write_experiment_report(
    path: str | Path,
    summary_rows: list[dict[str, object]],
    threshold_rows: list[dict[str, object]],
    distribution_rows: list[dict[str, object]],
    detail_rows: list[dict[str, object]],
    audit_rows: list[dict[str, object]] | None = None,
) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.suffix.lower() == ".xlsx":
        try:
            from openpyxl import Workbook
        except Exception:
            _write_csv(output.with_suffix(".csv"), summary_rows)
            return

        wb = Workbook()
        sheets = [
            ("採用監査", audit_rows or []),
            ("戦略比較", summary_rows),
            ("閾値別", threshold_rows),
            ("可変RR内訳", distribution_rows),
            ("候補", detail_rows),
        ]
        for index, (title, rows) in enumerate(sheets):
            ws = wb.active if index == 0 else wb.create_sheet(title)
            ws.title = title
            _append_rows(ws, rows)
            _format_sheet(ws)
        wb.save(output)
    else:
        _write_csv(output, summary_rows)


def _choose_setup_ladder(choices: dict[float, Candidate], rr_values: list[float]) -> Candidate | None:
    low = rr_values[0]
    mid = rr_values[len(rr_values) // 2]
    high = rr_values[-1]
    high_candidate = choices.get(high)
    mid_candidate = choices.get(mid)
    if high_candidate is not None and setup_quality_score(high_candidate) >= 42.0:
        return high_candidate
    if mid_candidate is not None and setup_quality_score(mid_candidate) >= 35.0:
        return mid_candidate
    return choices.get(low) or mid_candidate or high_candidate


def _choose_space_ladder(choices: dict[float, Candidate], rr_values: list[float]) -> Candidate | None:
    low = rr_values[0]
    mid = rr_values[len(rr_values) // 2]
    high = rr_values[-1]
    high_candidate = choices.get(high)
    mid_candidate = choices.get(mid)
    if high_candidate is not None and setup_quality_score(high_candidate) >= 40.0 and _tp_obstacles(high_candidate) == 0:
        return high_candidate
    if mid_candidate is not None and setup_quality_score(mid_candidate) >= 34.0 and _tp_obstacles(mid_candidate) <= 1:
        return mid_candidate
    return choices.get(low) or mid_candidate or high_candidate


def _choose_side_ladder(choices: dict[float, Candidate], rr_values: list[float]) -> Candidate | None:
    low = rr_values[0]
    mid = rr_values[len(rr_values) // 2]
    high = rr_values[-1]
    sample = next(iter(choices.values()), None)
    if sample is None:
        return None

    low_candidate = choices.get(low)
    mid_candidate = choices.get(mid)
    high_candidate = choices.get(high)
    if sample.side == "buy":
        if high_candidate is not None and _buy_can_try_high_rr(high_candidate):
            return high_candidate
        if mid_candidate is not None and setup_quality_score(mid_candidate) >= 34.0:
            return mid_candidate
        return low_candidate or mid_candidate or high_candidate

    if high_candidate is not None:
        return high_candidate
    if mid_candidate is not None and setup_quality_score(mid_candidate) >= 34.0:
        return mid_candidate
    return low_candidate or mid_candidate or high_candidate


def _buy_can_try_high_rr(candidate: Candidate) -> bool:
    return (
        setup_quality_score(candidate) >= 48.0
        and _tp_obstacles(candidate) == 0
        and _feature_number(candidate, "m15_extension_atr") <= 1.2
        and _feature_number(candidate, "M15_rsi14") <= 62.0
    )


def _sell_can_try_high_rr(candidate: Candidate) -> bool:
    return (
        setup_quality_score(candidate) >= 38.0
        and _tp_obstacles(candidate) <= 2
        and _feature_number(candidate, "m15_extension_atr") <= 2.0
        and _feature_number(candidate, "M15_rsi14", default=50.0) >= 34.0
    )


def _summary_row(
    strategy: str,
    policy: str,
    rr: float | str,
    side: str,
    min_score: float | None,
    candidates: list[Candidate],
    results: list[BacktestResult],
) -> dict[str, object]:
    summary = summarize(candidates, results)["overall"]
    return {
        "strategy": strategy,
        "policy": policy,
        "rr": rr,
        "side": side,
        "min_score": "" if min_score is None else min_score,
        **summary,
    }


def _detail_rows(strategy: str, candidates: list[Candidate], results: list[BacktestResult]) -> list[dict[str, object]]:
    rows = candidate_result_rows(candidates, results)
    for row in rows:
        row["strategy"] = strategy
        row["selected_rr"] = row.get("risk_reward", "")
    return rows


def _rr_distribution_rows(strategy: str, policy: str, candidates: list[Candidate]) -> list[dict[str, object]]:
    counts = Counter(candidate.risk_reward for candidate in candidates)
    total = len(candidates) or 1
    return [
        {
            "strategy": strategy,
            "policy": policy,
            "selected_rr": rr,
            "count": count,
            "share": round(count / total, 4),
        }
        for rr, count in sorted(counts.items())
    ]


def _filter_side(candidates: list[Candidate], side: str) -> list[Candidate]:
    if side == "both":
        return candidates
    return [candidate for candidate in candidates if candidate.side == side]


def _filter_min_score(candidates: list[Candidate], min_score: float | None) -> list[Candidate]:
    if min_score is None:
        return candidates
    return [candidate for candidate in candidates if candidate.score >= min_score]


def _balance_score(row: dict[str, object]) -> float:
    avg_r = _to_float(row.get("avg_r"))
    pf = min(_to_float(row.get("pf")), 10.0)
    max_drawdown = _to_float(row.get("max_drawdown_r"))
    return avg_r + (pf * 0.05) - (max_drawdown * 0.01)


def _group_detail_rows(rows: list[dict[str, object]], key: str) -> dict[str, list[dict[str, object]]]:
    grouped: dict[str, list[dict[str, object]]] = {}
    for row in rows:
        group = str(row.get(key) or "")
        grouped.setdefault(group, []).append(row)
    return grouped


def _side_summary(side: str, rows: list[dict[str, object]]) -> dict[str, object]:
    net_values = [_to_float(row.get("net_r_multiple")) for row in rows]
    gross_profit = sum(value for value in net_values if value > 0)
    gross_loss = -sum(value for value in net_values if value < 0)
    wins = sum(1 for row in rows if row.get("result") == "win")
    losses = sum(1 for row in rows if row.get("result") == "loss")
    total = len(rows)
    return {
        "side": side,
        "count": total,
        "wins": wins,
        "losses": losses,
        "win_rate": round(wins / total, 4) if total else 0.0,
        "avg_r": round(sum(net_values) / total, 4) if total else 0.0,
        "pf": round(gross_profit / gross_loss, 4) if gross_loss > 0 else 0.0,
        "total_r": round(sum(net_values), 4),
    }


def _to_float(value: object) -> float:
    if value is None or isinstance(value, bool):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _to_int(value: object) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


def _tp_obstacles(candidate: Candidate) -> int:
    return max(
        _feature_int(candidate, "tp_obstacle_count"),
        _feature_int(candidate, "M5_tp_obstacle_count"),
        _feature_int(candidate, "M15_tp_obstacle_count"),
    )


def _feature_int(candidate: Candidate, name: str) -> int:
    value = candidate.features.get(name)
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _feature_number(candidate: Candidate, name: str, default: float = 0.0) -> float:
    value = candidate.features.get(name)
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _num(value: object) -> float:
    if value is None or isinstance(value, bool):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _append_rows(ws, rows: list[dict[str, object]]) -> None:
    if not rows:
        return
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    ws.append(headers)
    for row in rows:
        ws.append([_excel_value(row.get(header)) for header in headers])


def _excel_value(value: object) -> object:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def _format_sheet(ws) -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
    ws.freeze_panes = "A2"
    for col in range(1, ws.max_column + 1):
        width = max(len(str(ws.cell(row, col).value or "")) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), 36)
    if ws.max_row > 1 and ws.max_column > 1:
        safe_title = "".join(ch if ch.isascii() and ch.isalnum() else "_" for ch in ws.title)
        if not any(ch.isalnum() for ch in safe_title):
            safe_title = f"Sheet{ws.parent.worksheets.index(ws) + 1}"
        table = Table(displayName=f"Table_{safe_title}", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)


def _write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def _append_markdown_table(lines: list[str], rows: list[dict[str, object]]) -> None:
    if not rows:
        lines.append("_No rows._")
        return
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    lines.append("| " + " | ".join(_md(header) for header in headers) + " |")
    lines.append("|" + "|".join("---" for _ in headers) + "|")
    for row in rows:
        lines.append("| " + " | ".join(_md(row.get(header, "")) for header in headers) + " |")


def _md(value: object) -> str:
    if isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False, sort_keys=True)
    else:
        text = str(value)
    return text.replace("|", "\\|").replace("\r", " ").replace("\n", " ")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare fixed and variable RR swing strategies.")
    parser.add_argument("--history", default="runtime/latest_history_168h.json")
    parser.add_argument("--output", default="reports/rr_strategy_experiment.xlsx")
    parser.add_argument("--output-json", default="", help="Optional JSON report with strategy summaries and adoption audit.")
    parser.add_argument("--output-md", default="", help="Optional Markdown report with strategy summaries and adoption audit.")
    parser.add_argument("--rr-values", type=parse_rr_values, default=list(DEFAULT_RR_VALUES))
    parser.add_argument("--variable-policies", type=parse_policies, default=parse_policies(",".join(VARIABLE_POLICIES)))
    parser.add_argument("--min-score", type=float, default=None)
    parser.add_argument("--side", choices=("buy", "sell", "both"), default="both")
    parser.add_argument("--max-hold-minutes", type=int, default=60)
    parser.add_argument("--swing-left", type=int, default=3)
    parser.add_argument("--swing-right", type=int, default=3)
    parser.add_argument("--min-atr-distance", type=float, default=0.5)
    parser.add_argument("--max-risk-atr", type=float, default=3.0)
    parser.add_argument("--score-profile", choices=("side", "balanced", "buy", "sell"), default="side")
    parser.add_argument("--include-blackout-times", action="store_true", help="Include rollover/news-proxy no-entry windows in candidate generation.")
    parser.add_argument("--calendar", default="runtime/economic_calendar.json", help="Optional economic calendar JSON/CSV in MT5 server time.")
    parser.add_argument("--calendar-input-utc-offset", type=float, default=None, help="UTC offset of naive calendar times, e.g. 9 for JST. Omit when calendar is already MT5 server time.")
    parser.add_argument("--calendar-server-utc-offset", type=float, default=None, help="MT5 server UTC offset used when converting calendar times.")
    parser.add_argument("--news-before-minutes", type=int, default=10)
    parser.add_argument("--news-after-minutes", type=int, default=10)
    parser.add_argument("--news-min-impact", default="high", choices=("low", "medium", "high"))
    parser.add_argument("--news-currencies", default="USD,XAU,ALL")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    history = load_history(args.history)
    calendar_events = load_economic_calendar(
        args.calendar,
        input_utc_offset_hours=args.calendar_input_utc_offset,
        server_utc_offset_hours=args.calendar_server_utc_offset,
    )
    summary_rows, threshold_rows, distribution_rows, detail_rows = run_rr_experiment(
        history,
        rr_values=args.rr_values,
        side=args.side,
        min_score=args.min_score,
        max_hold_minutes=args.max_hold_minutes,
        swing_left=args.swing_left,
        swing_right=args.swing_right,
        min_atr_distance=args.min_atr_distance,
        max_risk_atr=args.max_risk_atr,
        score_profile=args.score_profile,
        exclude_blackout_times=not args.include_blackout_times,
        blackout_events=calendar_events,
        news_before_minutes=args.news_before_minutes,
        news_after_minutes=args.news_after_minutes,
        news_min_impact=args.news_min_impact,
        news_currencies=parse_currencies(args.news_currencies),
        variable_policies=args.variable_policies,
    )
    audit_rows = rr_strategy_adoption_audit(summary_rows, threshold_rows, detail_rows)
    metadata = experiment_metadata(args, history, calendar_events)
    write_experiment_report(args.output, summary_rows, threshold_rows, distribution_rows, detail_rows, audit_rows)
    if args.output_json:
        write_experiment_json(
            args.output_json,
            summary_rows,
            threshold_rows,
            distribution_rows,
            detail_rows,
            audit_rows,
            metadata=metadata,
        )
    if args.output_md:
        write_experiment_markdown(
            args.output_md,
            summary_rows,
            threshold_rows,
            distribution_rows,
            audit_rows,
            metadata=metadata,
        )
    for row in summary_rows:
        print(row)
    print(f"wrote {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
