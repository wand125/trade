from __future__ import annotations

import csv
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from analysis.diagnostics import feature_diagnostics, max_drawdown, max_losing_streak, threshold_diagnostics
from analysis.models import BacktestResult, Candidate


def candidate_result_rows(candidates: list[Candidate], results: list[BacktestResult]) -> list[dict[str, object]]:
    result_by_id = {result.candidate_id: result for result in results}
    rows: list[dict[str, object]] = []
    for candidate in candidates:
        result = result_by_id.get(candidate.candidate_id)
        row: dict[str, object] = {
            "candidate_id": candidate.candidate_id,
            "time": candidate.time_text,
            "session": _session(candidate.time_text),
            "symbol": candidate.symbol,
            "side": candidate.side,
            "pattern": candidate.pattern,
            "entry": round(candidate.entry, 2),
            "sl": round(candidate.sl, 2),
            "tp": round(candidate.tp, 2),
            "risk": round(candidate.risk, 2),
            "risk_reward": round(candidate.risk_reward, 2),
            "score": round(candidate.score, 2),
            "swing_time": candidate.swing_time,
            "swing_price": round(candidate.swing_price, 2),
        }
        row.update(candidate.score_parts)
        if result is not None:
            row.update(
                {
                    "result": result.result,
                    "r_multiple": round(result.r_multiple, 3),
                    "net_r_multiple": round(result.net_r_multiple, 3),
                    "exit_time": result.exit_time,
                    "exit_price": round(result.exit_price, 2),
                    "exit_reason": result.exit_reason,
                    "bars_held": result.bars_held,
                }
            )
        for name, value in sorted(candidate.features.items()):
            if name in row:
                row[f"feature_{name}"] = value
            else:
                row[name] = value
        rows.append(row)
    return rows


def summarize(candidates: list[Candidate], results: list[BacktestResult]) -> dict[str, object]:
    rows = candidate_result_rows(candidates, results)
    return {
        "overall": _summary(rows),
        "score_bands": _group_summary(rows, "score_band"),
        "side": _group_summary(rows, "side"),
        "session": _group_summary(rows, "session"),
        "pattern": _group_summary(rows, "pattern"),
    }


def print_summary(summary: dict[str, object]) -> None:
    overall = summary["overall"]
    print("overall", overall)
    print("score_bands")
    for row in summary["score_bands"]:
        print(row)
    print("side")
    for row in summary["side"]:
        print(row)


def write_report(path: str | Path, candidates: list[Candidate], results: list[BacktestResult]) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    rows = candidate_result_rows(candidates, results)
    if output.suffix.lower() == ".xlsx":
        _write_xlsx(output, rows, summarize(candidates, results), candidates, results)
    elif output.suffix.lower() == ".md":
        _write_markdown(output, candidates, results)
    else:
        _write_csv(output, rows)


def _write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=_row_headers(rows))
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(
    path: Path,
    rows: list[dict[str, object]],
    summary: dict[str, object],
    candidates: list[Candidate],
    results: list[BacktestResult],
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except Exception:
        _write_csv(path.with_suffix(".csv"), rows)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "候補"
    if rows:
        headers = _row_headers(rows)
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header) for header in headers])
        _format_sheet(ws)
        table = Table(displayName="Candidates", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)

    ws_summary = wb.create_sheet("サマリー")
    _append_mapping(ws_summary, "overall", summary["overall"])
    ws_bands = wb.create_sheet("スコア帯")
    _append_rows(ws_bands, summary["score_bands"])
    ws_side = wb.create_sheet("方向別")
    _append_rows(ws_side, summary["side"])
    ws_session = wb.create_sheet("時間帯別")
    _append_rows(ws_session, summary["session"])
    ws_pattern = wb.create_sheet("パターン別")
    _append_rows(ws_pattern, summary["pattern"])
    ws_thresholds = wb.create_sheet("閾値別")
    _append_rows(ws_thresholds, threshold_diagnostics(candidates, results))
    ws_features = wb.create_sheet("特徴量診断")
    _append_rows(ws_features, feature_diagnostics(candidates, results))
    for sheet in wb.worksheets:
        _format_sheet(sheet)
    wb.save(path)


def _write_markdown(path: Path, candidates: list[Candidate], results: list[BacktestResult]) -> None:
    path.write_text(format_markdown_report(candidates, results), encoding="utf-8")


def format_markdown_report(candidates: list[Candidate], results: list[BacktestResult]) -> str:
    summary = summarize(candidates, results)
    thresholds = threshold_diagnostics(candidates, results)
    lines = [
        "# Signal Score Backtest Report",
        "",
        "## Overall",
        "",
    ]
    _append_mapping_table(lines, summary["overall"])
    lines.extend(["", "## Score Bands", ""])
    _append_markdown_table(lines, summary["score_bands"])
    lines.extend(["", "## Side", ""])
    _append_markdown_table(lines, summary["side"])
    lines.extend(["", "## Session", ""])
    _append_markdown_table(lines, summary["session"])
    lines.extend(["", "## Pattern", ""])
    _append_markdown_table(lines, summary["pattern"])
    lines.extend(["", "## Thresholds", ""])
    _append_markdown_table(lines, thresholds)
    return "\n".join(lines) + "\n"


def _append_mapping_table(lines: list[str], mapping: dict[str, object]) -> None:
    lines.append("| metric | value |")
    lines.append("|---|---:|")
    for key, value in mapping.items():
        lines.append(f"| {_md(key)} | {_md(value)} |")


def _append_markdown_table(lines: list[str], rows: list[dict[str, object]]) -> None:
    if not rows:
        lines.append("_No rows._")
        return
    headers = _row_headers(rows)
    lines.append("| " + " | ".join(_md(header) for header in headers) + " |")
    lines.append("|" + "|".join("---" for _ in headers) + "|")
    for row in rows:
        lines.append("| " + " | ".join(_md(row.get(header, "")) for header in headers) + " |")


def _md(value: object) -> str:
    text = str(value)
    text = text.replace("|", "\\|").replace("\r", " ").replace("\n", " ")
    return text


def _append_mapping(ws, title: str, mapping: dict[str, object]) -> None:
    ws.append([title, "value"])
    for key, value in mapping.items():
        ws.append([key, value])


def _append_rows(ws, rows: list[dict[str, object]]) -> None:
    if not rows:
        return
    headers = _row_headers(rows)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])


def _row_headers(rows: list[dict[str, object]]) -> list[str]:
    headers: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row.keys():
            if key in seen:
                continue
            seen.add(key)
            headers.append(key)
    return headers


def _format_sheet(ws) -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
    ws.freeze_panes = "A2"
    for col in range(1, ws.max_column + 1):
        width = max(len(str(ws.cell(row, col).value or "")) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), 32)


def _group_summary(rows: list[dict[str, object]], key: str) -> list[dict[str, object]]:
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        group_key = _score_band(row["score"]) if key == "score_band" else str(row.get(key, ""))
        grouped[group_key].append(row)
    return [{"group": group, **_summary(group_rows)} for group, group_rows in sorted(grouped.items())]


def _summary(rows: list[dict[str, object]]) -> dict[str, object]:
    total = len(rows)
    if total == 0:
        return {
            "count": 0,
            "wins": 0,
            "losses": 0,
            "timeouts": 0,
            "win_rate": 0.0,
            "avg_r": 0.0,
            "pf": 0.0,
            "total_r": 0.0,
            "max_losing_streak": 0,
            "max_drawdown_r": 0.0,
            "expectancy_r": 0.0,
        }
    wins = [row for row in rows if row.get("result") == "win"]
    losses = [row for row in rows if row.get("result") == "loss"]
    timeouts = [row for row in rows if row.get("result") == "timeout"]
    net_values = [float(row.get("net_r_multiple", 0.0) or 0.0) for row in rows]
    ordered_values = [
        float(row.get("net_r_multiple", 0.0) or 0.0)
        for row in sorted(rows, key=lambda item: str(item.get("time", "")))
    ]
    ordered_results = [
        BacktestResult(
            candidate_id=str(row.get("candidate_id", "")),
            result=str(row.get("result", "")),
            r_multiple=float(row.get("r_multiple", 0.0) or 0.0),
            net_r_multiple=float(row.get("net_r_multiple", 0.0) or 0.0),
            exit_time=str(row.get("exit_time", "")),
            exit_price=float(row.get("exit_price", 0.0) or 0.0),
            exit_reason=str(row.get("exit_reason", "")),
            bars_held=int(row.get("bars_held", 0) or 0),
        )
        for row in sorted(rows, key=lambda item: str(item.get("time", "")))
    ]
    gross_profit = sum(value for value in net_values if value > 0)
    gross_loss = -sum(value for value in net_values if value < 0)
    return {
        "count": total,
        "wins": len(wins),
        "losses": len(losses),
        "timeouts": len(timeouts),
        "win_rate": round(len(wins) / total, 4),
        "avg_r": round(sum(net_values) / total, 4),
        "pf": round(gross_profit / gross_loss, 4) if gross_loss > 0 else 0.0,
        "total_r": round(sum(net_values), 4),
        "max_losing_streak": max_losing_streak(ordered_results),
        "max_drawdown_r": round(max_drawdown(ordered_values), 4),
        "expectancy_r": round(sum(net_values) / total, 4),
    }


def _score_band(score: object) -> str:
    value = float(score)
    lower = int(value // 10 * 10)
    upper = lower + 10
    return f"{lower}-{upper}"


def _session(time_text: str) -> str:
    try:
        hour = datetime.strptime(time_text, "%Y.%m.%d %H:%M").hour
    except ValueError:
        return "unknown"
    if hour in (0, 23):
        return "rollover"
    if 7 <= hour < 15:
        return "tokyo"
    if 15 <= hour < 21:
        return "london"
    if 21 <= hour or hour < 6:
        return "new_york"
    return "quiet"
