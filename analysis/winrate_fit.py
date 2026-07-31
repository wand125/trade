from __future__ import annotations

import argparse
import csv
import json
import sys
from dataclasses import dataclass
from datetime import timedelta
from datetime import datetime
from pathlib import Path
from statistics import mean, median

if __package__ is None or __package__ == "":
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from analysis.backtest import run_backtest
from analysis.candidate_generator import generate_candidates
from analysis.economic_calendar import load_economic_calendar, parse_currencies
from analysis.market_data import load_history
from analysis.models import BacktestResult, Candidate
from analysis.reports import candidate_result_rows
from analysis.market_data import TIME_FORMAT


@dataclass(frozen=True)
class FitRule:
    feature: str
    op: str
    threshold: float

    def passes(self, record: "FitRecord") -> bool:
        value = record.features.get(self.feature)
        if value is None:
            return False
        if self.op == ">=":
            return value >= self.threshold
        return value <= self.threshold

    def expression(self) -> str:
        return f"{self.feature} {self.op} {self.threshold:.5g}"


@dataclass(frozen=True)
class FitRecord:
    candidate: Candidate
    result: BacktestResult
    features: dict[str, float]

    @property
    def is_win(self) -> bool:
        return self.result.result == "win"


def build_records(candidates: list[Candidate], results: list[BacktestResult]) -> list[FitRecord]:
    result_by_id = {result.candidate_id: result for result in results}
    records: list[FitRecord] = []
    for candidate in candidates:
        result = result_by_id.get(candidate.candidate_id)
        if result is None:
            continue
        records.append(FitRecord(candidate=candidate, result=result, features=_numeric_features(candidate)))
    return sorted(records, key=lambda record: record.candidate.time)


def fit_winrate_rules(
    records: list[FitRecord],
    *,
    train_ratio: float = 0.7,
    purge_records: int = 0,
    embargo_records: int = 0,
    embargo_minutes: int = 0,
    min_count: int = 20,
    max_rules: int = 3,
    min_avg_r: float = 0.0,
    quantiles: tuple[float, ...] = (0.15, 0.25, 0.35, 0.5, 0.65, 0.75, 0.85),
) -> tuple[list[FitRule], list[dict[str, object]], list[dict[str, object]], list[FitRecord], list[FitRecord]]:
    train, test = split_records(
        records,
        train_ratio,
        purge_records=purge_records,
        embargo_records=embargo_records,
        embargo_minutes=embargo_minutes,
    )
    selected: list[FitRule] = []
    step_rows: list[dict[str, object]] = []
    search_rows: list[dict[str, object]] = []
    current_train = train
    baseline = metrics(current_train)
    step_rows.append({"step": 0, "rule": "baseline", "rules": "", **baseline})

    for step in range(1, max_rules + 1):
        candidates = candidate_rules(current_train, quantiles=quantiles)
        scored: list[tuple[tuple[float, float, int], FitRule, list[FitRecord], dict[str, object]]] = []
        for rule in candidates:
            if rule in selected:
                continue
            filtered = apply_rules(current_train, [rule])
            if len(filtered) < min_count:
                continue
            row = metrics(filtered)
            if float(row["avg_r"]) < min_avg_r:
                continue
            key = (float(row["win_rate"]), float(row["avg_r"]), int(row["count"]))
            scored.append((key, rule, filtered, row))
        if not scored:
            break
        scored.sort(key=lambda item: item[0], reverse=True)
        for rank, (_, rule, _, row) in enumerate(scored[:50], start=1):
            search_rows.append({"step": step, "rank": rank, "rule": rule.expression(), **row})

        _, best_rule, filtered_train, best_row = scored[0]
        if float(best_row["win_rate"]) <= float(metrics(current_train)["win_rate"]):
            break
        selected.append(best_rule)
        current_train = filtered_train
        step_rows.append(
            {
                "step": step,
                "rule": best_rule.expression(),
                "rules": " AND ".join(rule.expression() for rule in selected),
                **best_row,
            }
        )

    return selected, step_rows, search_rows, train, test


def walk_forward_fit(
    records: list[FitRecord],
    *,
    folds: int = 4,
    train_window: int = 40,
    test_window: int = 12,
    purge_records: int = 0,
    embargo_records: int = 0,
    embargo_minutes: int = 0,
    min_count: int = 12,
    max_rules: int = 3,
    min_avg_r: float = 0.0,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    windows = walk_forward_windows(
        records,
        folds=folds,
        train_window=train_window,
        test_window=test_window,
        purge_records=purge_records,
        embargo_records=embargo_records,
        embargo_minutes=embargo_minutes,
    )
    fold_rows: list[dict[str, object]] = []
    rule_rows: list[dict[str, object]] = []
    for fold_index, (train, test) in enumerate(windows, start=1):
        rules, _, search_rows = robust_fit_rules(
            train,
            min_count=min(min_count, max(1, len(train) // 3)),
            max_rules=max_rules,
            min_avg_r=min_avg_r,
        )
        train_fitted = apply_rules(train, rules)
        test_fitted = apply_rules(test, rules)
        rule_text = " AND ".join(rule.expression() for rule in rules)
        fold_rows.append(
            {
                "fold": fold_index,
                "train_start": train[0].candidate.time_text if train else "",
                "train_end": train[-1].candidate.time_text if train else "",
                "test_start": test[0].candidate.time_text if test else "",
                "test_end": test[-1].candidate.time_text if test else "",
                "rules": rule_text,
                "purge_records": purge_records,
                "embargo_records": embargo_records,
                "embargo_minutes": embargo_minutes,
                **_prefixed_metrics("train_", metrics(train)),
                **_prefixed_metrics("train_fitted_", metrics(train_fitted)),
                **_prefixed_metrics("test_", metrics(test)),
                **_prefixed_metrics("test_fitted_", metrics(test_fitted)),
            }
        )
        for row in search_rows[:20]:
            rule_rows.append({"fold": fold_index, **row})
    fold_rows.append(aggregate_walk_forward_row(fold_rows))
    return fold_rows, rule_rows


def robust_fit_rules(
    records: list[FitRecord],
    *,
    validation_folds: int = 3,
    purge_records: int = 0,
    embargo_records: int = 0,
    embargo_minutes: int = 0,
    min_count: int = 12,
    max_rules: int = 3,
    min_avg_r: float = 0.0,
    candidate_keep: int = 40,
) -> tuple[list[FitRule], list[dict[str, object]], list[dict[str, object]]]:
    selected: list[FitRule] = []
    step_rows: list[dict[str, object]] = [{"step": 0, "rule": "baseline", "rules": "", **metrics(records)}]
    search_rows: list[dict[str, object]] = []
    baseline = robust_rule_set_score(
        records,
        [],
        folds=validation_folds,
        purge_records=purge_records,
        embargo_records=embargo_records,
        embargo_minutes=embargo_minutes,
        min_count=min_count,
    )
    current_key = robust_score_key(baseline)

    for step in range(1, max_rules + 1):
        remaining = [rule for rule in candidate_rules(threshold_source(records), quantiles=(0.2, 0.35, 0.5, 0.65, 0.8)) if rule not in selected]
        single_scores: list[tuple[tuple[float, float, float, int], FitRule, dict[str, object]]] = []
        for rule in remaining:
            row = robust_rule_set_score(
                records,
                selected + [rule],
                folds=validation_folds,
                purge_records=purge_records,
                embargo_records=embargo_records,
                embargo_minutes=embargo_minutes,
                min_count=max(1, min_count // 2),
            )
            if int(row["total_validation_count"]) < max(1, min_count // 2):
                continue
            if float(row["mean_validation_avg_r"]) < min_avg_r:
                continue
            single_scores.append((robust_score_key(row), rule, row))
        single_scores.sort(key=lambda item: item[0], reverse=True)

        scored_rule_sets: list[tuple[tuple[float, float, float, int], list[FitRule], dict[str, object]]] = []
        for _, rule, row in single_scores[:candidate_keep]:
            scored_rule_sets.append((robust_score_key(row), selected + [rule], row))

        if not scored_rule_sets:
            break
        scored_rule_sets.sort(key=lambda item: item[0], reverse=True)
        for rank, (_, rules, row) in enumerate(scored_rule_sets[:50], start=1):
            search_rows.append({"step": step, "rank": rank, "rules": " AND ".join(rule.expression() for rule in rules), **row})

        chosen: tuple[tuple[float, float, float, int], list[FitRule], dict[str, object]] | None = None
        current_metrics = metrics(apply_rules(records, selected))
        for candidate_key, candidate_ruleset, candidate_row in scored_rule_sets:
            fitted_records = apply_rules(records, candidate_ruleset)
            fitted_metrics = metrics(fitted_records)
            if int(fitted_metrics["count"]) < max(1, min_count):
                continue
            if float(fitted_metrics["win_rate"]) <= float(current_metrics["win_rate"]):
                continue
            chosen = (candidate_key, candidate_ruleset, candidate_row)
            break

        if chosen is None:
            break

        best_key, best_rules, best_row = chosen
        if best_key <= current_key:
            break
        selected = best_rules[:max_rules]
        current_key = best_key
        fitted_records = apply_rules(records, selected)
        step_rows.append(
            {
                "step": step,
                "rule": selected[-1].expression(),
                "rules": " AND ".join(rule.expression() for rule in selected),
                **metrics(fitted_records),
                **best_row,
            }
        )
        if len(selected) >= max_rules:
            break

    return selected, step_rows, search_rows


def robust_rule_set_score(
    records: list[FitRecord],
    rules: list[FitRule],
    *,
    folds: int,
    purge_records: int = 0,
    embargo_records: int = 0,
    embargo_minutes: int = 0,
    min_count: int,
) -> dict[str, object]:
    if len(records) < 6:
        validation_sets = [records]
    else:
        train_window = max(3, int(len(records) * 0.55))
        test_window = max(2, int(len(records) * 0.15))
        validation_sets = [
            test
            for _, test in walk_forward_windows(
                records,
                folds=folds,
                train_window=train_window,
                test_window=test_window,
                purge_records=purge_records,
                embargo_records=embargo_records,
                embargo_minutes=embargo_minutes,
            )
        ]
    rows: list[dict[str, object]] = []
    counts: list[int] = []
    for validation in validation_sets:
        fitted = apply_rules(validation, rules)
        row = metrics(fitted)
        rows.append(row)
        counts.append(int(row["count"]))
    rows_with_trades = [row for row in rows if int(row["count"]) > 0]
    win_rates = [float(row["win_rate"]) for row in rows_with_trades]
    avg_rs = [float(row["avg_r"]) for row in rows_with_trades]
    pfs = [float(row["pf"]) for row in rows_with_trades]
    total_count = sum(counts)
    return {
        "validation_folds": len(validation_sets),
        "folds_with_trades": sum(1 for count in counts if count > 0),
        "total_validation_count": total_count,
        "min_validation_count": min(counts) if counts else 0,
        "median_validation_win_rate": round(median(win_rates), 4) if win_rates else 0.0,
        "mean_validation_win_rate": round(mean(win_rates), 4) if win_rates else 0.0,
        "median_validation_avg_r": round(median(avg_rs), 4) if avg_rs else 0.0,
        "mean_validation_avg_r": round(mean(avg_rs), 4) if avg_rs else 0.0,
        "median_validation_pf": round(median(pfs), 4) if pfs else 0.0,
        "passes_min_count": total_count >= min_count,
    }


def robust_score_key(row: dict[str, object]) -> tuple[float, float, float, int]:
    count = int(row.get("total_validation_count", 0) or 0)
    folds_with_trades = int(row.get("folds_with_trades", 0) or 0)
    min_count = int(row.get("min_validation_count", 0) or 0)
    count_penalty = 0.0
    if count < 12:
        count_penalty += (12 - count) / 12 * 0.2
    if min_count < 2:
        count_penalty += 0.2
    win_rate_score = (
        float(row.get("median_validation_win_rate", 0.0) or 0.0)
        + float(row.get("mean_validation_win_rate", 0.0) or 0.0)
    ) / 2
    return (
        win_rate_score - count_penalty,
        float(row.get("median_validation_avg_r", 0.0) or 0.0),
        float(row.get("mean_validation_avg_r", 0.0) or 0.0),
        folds_with_trades,
    )


def threshold_source(records: list[FitRecord]) -> list[FitRecord]:
    if len(records) < 5:
        return records
    cutoff = max(2, int(len(records) * 0.7))
    return records[:cutoff]


def walk_forward_windows(
    records: list[FitRecord],
    *,
    folds: int,
    train_window: int,
    test_window: int,
    purge_records: int = 0,
    embargo_records: int = 0,
    embargo_minutes: int = 0,
) -> list[tuple[list[FitRecord], list[FitRecord]]]:
    if len(records) < 4:
        return []
    train_window = max(2, min(train_window, len(records) - 1))
    purge_records = max(0, purge_records)
    embargo_records = max(0, embargo_records)
    embargo_minutes = max(0, embargo_minutes)
    max_train_window = len(records) - embargo_records - 1
    if max_train_window < 2:
        return []
    train_window = max(2, min(train_window, max_train_window))
    test_window = max(1, min(test_window, len(records) - train_window - embargo_records))
    possible_starts = list(range(0, len(records) - train_window - embargo_records - test_window + 1))
    if not possible_starts:
        train, test = _apply_purge_embargo(
            records[:train_window],
            records[train_window + embargo_records :],
            purge_records=purge_records,
            embargo_minutes=embargo_minutes,
            split_train_end=records[train_window - 1],
            split_test_start=records[train_window] if train_window < len(records) else None,
        )
        return [(train, test)] if train and test else []
    if folds <= 1:
        starts = [possible_starts[-1]]
    else:
        starts = [
            possible_starts[round(i * (len(possible_starts) - 1) / (folds - 1))]
            for i in range(folds)
        ]
    windows: list[tuple[list[FitRecord], list[FitRecord]]] = []
    seen: set[int] = set()
    for start in starts:
        if start in seen:
            continue
        seen.add(start)
        train_end = start + train_window
        test_start = train_end + embargo_records
        train = records[start:train_end]
        test = records[test_start : test_start + test_window]
        train, test = _apply_purge_embargo(
            train,
            test,
            purge_records=purge_records,
            embargo_minutes=embargo_minutes,
            split_train_end=records[train_end - 1],
            split_test_start=records[train_end] if train_end < len(records) else None,
        )
        if train and test:
            windows.append((train, test))
    return windows


def _apply_purge_embargo(
    train: list[FitRecord],
    test: list[FitRecord],
    *,
    purge_records: int,
    embargo_minutes: int,
    split_train_end: FitRecord | None,
    split_test_start: FitRecord | None,
) -> tuple[list[FitRecord], list[FitRecord]]:
    if purge_records > 0:
        train = train[:-purge_records] if len(train) > purge_records else []
    if embargo_minutes <= 0:
        return train, test

    delta = timedelta(minutes=embargo_minutes)
    if split_test_start is not None:
        test_start_time = split_test_start.candidate.time
        train = [record for record in train if record.candidate.time + delta < test_start_time]
    if split_train_end is not None:
        train_end_time = split_train_end.candidate.time
        test = [record for record in test if record.candidate.time > train_end_time + delta]
    return train, test


def aggregate_walk_forward_row(rows: list[dict[str, object]]) -> dict[str, object]:
    actual_rows = [row for row in rows if row.get("fold") != "aggregate"]
    test_rates = [float(row.get("test_fitted_win_rate", 0.0) or 0.0) for row in actual_rows if int(row.get("test_fitted_count", 0) or 0) > 0]
    test_avg_r = [float(row.get("test_fitted_avg_r", 0.0) or 0.0) for row in actual_rows if int(row.get("test_fitted_count", 0) or 0) > 0]
    test_pf = [float(row.get("test_fitted_pf", 0.0) or 0.0) for row in actual_rows if int(row.get("test_fitted_count", 0) or 0) > 0]
    test_counts = [int(row.get("test_fitted_count", 0) or 0) for row in actual_rows]
    return {
        "fold": "aggregate",
        "train_start": "",
        "train_end": "",
        "test_start": "",
        "test_end": "",
        "rules": "",
        "folds": len(actual_rows),
        "folds_with_trades": sum(1 for count in test_counts if count > 0),
        "median_test_fitted_win_rate": round(median(test_rates), 4) if test_rates else 0.0,
        "mean_test_fitted_win_rate": round(mean(test_rates), 4) if test_rates else 0.0,
        "median_test_fitted_avg_r": round(median(test_avg_r), 4) if test_avg_r else 0.0,
        "mean_test_fitted_avg_r": round(mean(test_avg_r), 4) if test_avg_r else 0.0,
        "median_test_fitted_pf": round(median(test_pf), 4) if test_pf else 0.0,
        "mean_test_fitted_pf": round(mean(test_pf), 4) if test_pf else 0.0,
        "min_test_fitted_count": min(test_counts) if test_counts else 0,
        "total_test_fitted_count": sum(test_counts),
    }


def split_records(
    records: list[FitRecord],
    train_ratio: float,
    *,
    purge_records: int = 0,
    embargo_records: int = 0,
    embargo_minutes: int = 0,
) -> tuple[list[FitRecord], list[FitRecord]]:
    if not records:
        return [], []
    split_index = max(1, min(len(records) - 1, int(len(records) * train_ratio)))
    train = records[:split_index]
    test = records[split_index + max(0, embargo_records) :]
    return _apply_purge_embargo(
        train,
        test,
        purge_records=max(0, purge_records),
        embargo_minutes=max(0, embargo_minutes),
        split_train_end=records[split_index - 1],
        split_test_start=records[split_index],
    )


def apply_rules(records: list[FitRecord], rules: list[FitRule]) -> list[FitRecord]:
    if not rules:
        return records
    return [record for record in records if all(rule.passes(record) for rule in rules)]


def candidate_rules(records: list[FitRecord], *, quantiles: tuple[float, ...]) -> list[FitRule]:
    names = sorted({name for record in records for name in record.features if is_fit_feature_allowed(name)})
    rules: list[FitRule] = []
    for name in names:
        values = sorted(record.features[name] for record in records if name in record.features)
        if len(set(values)) < 2:
            continue
        for quantile in quantiles:
            threshold = values[min(len(values) - 1, max(0, int((len(values) - 1) * quantile)))]
            rules.append(FitRule(name, ">=", threshold))
            rules.append(FitRule(name, "<=", threshold))
    return rules


def is_fit_feature_allowed(name: str) -> bool:
    nonstationary_suffixes = (
        "_ema_fast",
        "_ema_slow",
        "_ema_mid",
        "_ema_long",
    )
    if name.endswith(nonstationary_suffixes):
        return False
    return name not in {"risk_reward"}


def metrics(records: list[FitRecord]) -> dict[str, object]:
    if not records:
        return {
            "count": 0,
            "wins": 0,
            "losses": 0,
            "timeouts": 0,
            "win_rate": 0.0,
            "avg_r": 0.0,
            "pf": 0.0,
            "total_r": 0.0,
            "avg_score": 0.0,
        }
    wins = sum(1 for record in records if record.result.result == "win")
    losses = sum(1 for record in records if record.result.result == "loss")
    timeouts = sum(1 for record in records if record.result.result == "timeout")
    net_values = [record.result.net_r_multiple for record in records]
    gross_profit = sum(value for value in net_values if value > 0)
    gross_loss = -sum(value for value in net_values if value < 0)
    return {
        "count": len(records),
        "wins": wins,
        "losses": losses,
        "timeouts": timeouts,
        "win_rate": round(wins / len(records), 4),
        "avg_r": round(sum(net_values) / len(records), 4),
        "pf": round(gross_profit / gross_loss, 4) if gross_loss > 0 else 0.0,
        "total_r": round(sum(net_values), 4),
        "avg_score": round(mean(record.candidate.score for record in records), 2),
    }


def report_rows(
    records: list[FitRecord],
    train: list[FitRecord],
    test: list[FitRecord],
    rules: list[FitRule],
) -> list[dict[str, object]]:
    return [
        {"dataset": "all_baseline", "rules": "", **metrics(records)},
        {"dataset": "train_baseline", "rules": "", **metrics(train)},
        {"dataset": "test_baseline", "rules": "", **metrics(test)},
        {"dataset": "train_fitted", "rules": " AND ".join(rule.expression() for rule in rules), **metrics(apply_rules(train, rules))},
        {"dataset": "test_fitted", "rules": " AND ".join(rule.expression() for rule in rules), **metrics(apply_rules(test, rules))},
        {"dataset": "all_fitted", "rules": " AND ".join(rule.expression() for rule in rules), **metrics(apply_rules(records, rules))},
    ]


def adoption_decision_row(
    summary_rows: list[dict[str, object]],
    rules: list[FitRule],
    *,
    min_test_count: int = 5,
    min_test_avg_r: float = 0.0,
    min_test_pf: float = 1.0,
) -> dict[str, object]:
    by_dataset = {str(row.get("dataset")): row for row in summary_rows}
    baseline = by_dataset.get("test_baseline", {})
    fitted = by_dataset.get("test_fitted", {})
    reasons: list[str] = []
    if not rules:
        reasons.append("no fitted rules")
    if int(fitted.get("count", 0) or 0) < min_test_count:
        reasons.append(f"test fitted count < {min_test_count}")
    if float(fitted.get("avg_r", 0.0) or 0.0) < min_test_avg_r:
        reasons.append(f"test fitted avg_r < {min_test_avg_r:g}")
    if float(fitted.get("pf", 0.0) or 0.0) < min_test_pf:
        reasons.append(f"test fitted pf < {min_test_pf:g}")
    if float(fitted.get("avg_r", 0.0) or 0.0) < float(baseline.get("avg_r", 0.0) or 0.0):
        reasons.append("test fitted avg_r below test baseline")
    return {
        "dataset": "adoption_decision",
        "rules": " AND ".join(rule.expression() for rule in rules),
        "adopted": not reasons,
        "reasons": "; ".join(reasons) if reasons else "passed final test gate",
        "min_test_count": min_test_count,
        "min_test_avg_r": min_test_avg_r,
        "min_test_pf": min_test_pf,
    }


def _prefixed_metrics(prefix: str, row: dict[str, object]) -> dict[str, object]:
    return {f"{prefix}{key}": value for key, value in row.items()}


def write_fit_report(
    path: str | Path,
    summary_rows: list[dict[str, object]],
    step_rows: list[dict[str, object]],
    search_rows: list[dict[str, object]],
    walk_rows: list[dict[str, object]],
    walk_rule_rows: list[dict[str, object]],
    selected_records: list[FitRecord],
) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    detail_rows = candidate_result_rows([record.candidate for record in selected_records], [record.result for record in selected_records])
    if output.suffix.lower() == ".xlsx":
        try:
            from openpyxl import Workbook
        except Exception:
            _write_csv(output.with_suffix(".csv"), summary_rows)
            return
        wb = Workbook()
        for index, (title, rows) in enumerate(
            (
                ("サマリー", summary_rows),
                ("採用ルール", step_rows),
                ("探索上位", search_rows),
                ("ウォークフォワード", walk_rows),
                ("WF探索上位", walk_rule_rows),
                ("通過候補", detail_rows),
            )
        ):
            ws = wb.active if index == 0 else wb.create_sheet(title)
            ws.title = title
            _append_rows(ws, rows)
            _format_sheet(ws, index)
        wb.save(output)
    else:
        _write_csv(output, summary_rows)


def write_fit_json(
    path: str | Path,
    *,
    summary_rows: list[dict[str, object]],
    step_rows: list[dict[str, object]],
    search_rows: list[dict[str, object]],
    walk_rows: list[dict[str, object]],
    walk_rule_rows: list[dict[str, object]],
    selected_records: list[FitRecord],
) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    adoption = next((row for row in summary_rows if row.get("dataset") == "adoption_decision"), {})
    payload = {
        "ok": True,
        "generated_at": datetime.now().strftime(TIME_FORMAT),
        "adoption_decision": adoption,
        "summary_rows": summary_rows,
        "step_rows": step_rows,
        "search_rows": search_rows,
        "walk_rows": walk_rows,
        "walk_rule_rows": walk_rule_rows,
        "selected_count": len(selected_records),
    }
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _numeric_features(candidate: Candidate) -> dict[str, float]:
    features: dict[str, float] = {"score": candidate.score, "risk": candidate.risk, "risk_reward": candidate.risk_reward}
    for name, value in candidate.score_parts.items():
        parsed = _to_float(value)
        if parsed is not None:
            features[name] = parsed
    for name, value in candidate.features.items():
        parsed = _to_float(value)
        if parsed is not None:
            features[name] = parsed
    return features


def _to_float(value: object) -> float | None:
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _append_rows(ws, rows: list[dict[str, object]]) -> None:
    if not rows:
        return
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])


def _format_sheet(ws, sheet_index: int = 0) -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
    ws.freeze_panes = "A2"
    for col in range(1, ws.max_column + 1):
        width = max(len(str(ws.cell(row, col).value or "")) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), 36)
    if ws.max_row > 1 and ws.max_column > 1:
        safe_title = "".join(ch if ch.isascii() and ch.isalnum() else "_" for ch in ws.title).strip("_") or "sheet"
        table = Table(displayName=f"Table_{sheet_index}_{safe_title}", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)


def _write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fit feature-threshold rules to improve win rate.")
    parser.add_argument("--history", default="runtime/latest_history_168h.json")
    parser.add_argument("--output", default="reports/winrate_fit.xlsx")
    parser.add_argument("--output-json", default="", help="Optional machine-readable fit report for promotion_gate.")
    parser.add_argument("--rr", type=float, default=4.0)
    parser.add_argument("--side", choices=("buy", "sell", "both"), default="both")
    parser.add_argument("--min-score", type=float, default=50.0)
    parser.add_argument("--train-ratio", type=float, default=0.7)
    parser.add_argument("--purge-records", type=int, default=0, help="Drop this many latest train records at train/test boundaries.")
    parser.add_argument("--embargo-records", type=int, default=0, help="Drop this many earliest test records after train windows.")
    parser.add_argument("--embargo-minutes", type=int, default=0, help="Drop records within this many minutes around train/test boundaries.")
    parser.add_argument("--min-count", type=int, default=20)
    parser.add_argument("--max-rules", type=int, default=3)
    parser.add_argument("--min-avg-r", type=float, default=0.0)
    parser.add_argument("--min-test-count", type=int, default=5)
    parser.add_argument("--min-test-avg-r", type=float, default=0.0)
    parser.add_argument("--min-test-pf", type=float, default=1.0)
    parser.add_argument("--validation-folds", type=int, default=3)
    parser.add_argument("--wf-folds", type=int, default=4)
    parser.add_argument("--wf-train-window", type=int, default=40)
    parser.add_argument("--wf-test-window", type=int, default=12)
    parser.add_argument("--max-hold-minutes", type=int, default=60)
    parser.add_argument("--score-profile", choices=("side", "balanced", "buy", "sell"), default="side")
    parser.add_argument("--include-blackout-times", action="store_true", help="Include rollover/news no-entry windows in candidate generation.")
    parser.add_argument("--calendar", default="runtime/economic_calendar.json", help="Optional economic calendar JSON/CSV in MT5 server time.")
    parser.add_argument("--calendar-input-utc-offset", type=float, default=None, help="UTC offset of naive calendar times, e.g. 9 for JST. Omit when calendar is already MT5 server time.")
    parser.add_argument("--calendar-server-utc-offset", type=float, default=None, help="MT5 server UTC offset used when converting calendar times.")
    parser.add_argument("--news-before-minutes", type=int, default=10)
    parser.add_argument("--news-after-minutes", type=int, default=10)
    parser.add_argument("--news-min-impact", default="high", choices=("low", "medium", "high"))
    parser.add_argument("--news-currencies", default="USD,XAU,ALL")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    history = load_history(args.history)
    calendar_events = load_economic_calendar(
        args.calendar,
        input_utc_offset_hours=args.calendar_input_utc_offset,
        server_utc_offset_hours=args.calendar_server_utc_offset,
    )
    candidates = generate_candidates(
        history,
        risk_reward=args.rr,
        min_score=args.min_score,
        score_profile=args.score_profile,
        exclude_blackout_times=not args.include_blackout_times,
        blackout_events=calendar_events,
        news_before_minutes=args.news_before_minutes,
        news_after_minutes=args.news_after_minutes,
        news_min_impact=args.news_min_impact,
        news_currencies=parse_currencies(args.news_currencies),
    )
    if args.side != "both":
        candidates = [candidate for candidate in candidates if candidate.side == args.side]
    results = run_backtest(candidates, history.bars("M1"), max_hold_minutes=args.max_hold_minutes, spread_price=history.spread_points * history.point)
    records = build_records(candidates, results)
    train, test = split_records(
        records,
        args.train_ratio,
        purge_records=args.purge_records,
        embargo_records=args.embargo_records,
        embargo_minutes=args.embargo_minutes,
    )
    rules, step_rows, search_rows = robust_fit_rules(
        train,
        validation_folds=args.validation_folds,
        purge_records=args.purge_records,
        embargo_records=args.embargo_records,
        embargo_minutes=args.embargo_minutes,
        min_count=max(3, min(args.min_count, len(train) // 3 if train else args.min_count)),
        max_rules=args.max_rules,
        min_avg_r=args.min_avg_r,
    )
    summary_rows = report_rows(records, train, test, rules)
    decision = adoption_decision_row(
        summary_rows,
        rules,
        min_test_count=args.min_test_count,
        min_test_avg_r=args.min_test_avg_r,
        min_test_pf=args.min_test_pf,
    )
    summary_rows.append(decision)
    adopted_rules = rules if decision["adopted"] else []
    selected_records = apply_rules(records, adopted_rules)
    walk_rows, walk_rule_rows = walk_forward_fit(
        records,
        folds=args.wf_folds,
        train_window=args.wf_train_window,
        test_window=args.wf_test_window,
        purge_records=args.purge_records,
        embargo_records=args.embargo_records,
        embargo_minutes=args.embargo_minutes,
        min_count=max(5, min(args.min_count, args.wf_train_window // 3)),
        max_rules=args.max_rules,
        min_avg_r=args.min_avg_r,
    )
    write_fit_report(args.output, summary_rows, step_rows, search_rows, walk_rows, walk_rule_rows, selected_records)
    if args.output_json:
        write_fit_json(
            args.output_json,
            summary_rows=summary_rows,
            step_rows=step_rows,
            search_rows=search_rows,
            walk_rows=walk_rows,
            walk_rule_rows=walk_rule_rows,
            selected_records=selected_records,
        )
    for row in summary_rows:
        print(row)
    for row in walk_rows:
        print(row)
    print(f"rules: {' AND '.join(rule.expression() for rule in rules) if rules else '(none)'}")
    print(f"adopted: {decision['adopted']} ({decision['reasons']})")
    print(f"wrote {args.output}")
    if args.output_json:
        print(f"wrote {args.output_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
