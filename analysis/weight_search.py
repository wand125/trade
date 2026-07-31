from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from statistics import mean, median

if __package__ is None or __package__ == "":
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from analysis.backtest import run_backtest
from analysis.candidate_generator import generate_candidates
from analysis.diagnostics import (
    DEFAULT_THRESHOLDS,
    ComponentWeights,
    component_weight_search,
    metrics_for_pairs,
    threshold_diagnostics,
    weighted_score,
)
from analysis.economic_calendar import load_economic_calendar, parse_currencies
from analysis.market_data import TIME_FORMAT, load_history
from analysis.models import BacktestResult, Candidate


def write_weight_search_report(
    path: str | Path,
    search_rows: list[dict[str, object]],
    baseline_rows: list[dict[str, object]],
    walk_forward_rows: list[dict[str, object]] | None = None,
    regime_rows: list[dict[str, object]] | None = None,
) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.suffix.lower() == ".xlsx":
        try:
            from openpyxl import Workbook
        except Exception:
            _write_csv(output.with_suffix(".csv"), search_rows)
            return
        wb = Workbook()
        sheets = [("重み探索", search_rows), ("現行閾値別", baseline_rows)]
        if walk_forward_rows is not None:
            sheets.append(("Walk Forward", walk_forward_rows))
        if regime_rows is not None:
            sheets.append(("Regime Search", regime_rows))
        for index, (title, rows) in enumerate(sheets):
            ws = wb.active if index == 0 else wb.create_sheet(title)
            ws.title = title
            _append_rows(ws, rows)
            _format_sheet(ws)
        wb.save(output)
    else:
        _write_csv(output, search_rows)


def build_weight_search_payload(
    *,
    settings: dict[str, object],
    candidate_count: int,
    result_count: int,
    search_rows: list[dict[str, object]],
    baseline_rows: list[dict[str, object]],
    max_rows: int,
    walk_forward: dict[str, object] | None = None,
    regime_search: dict[str, object] | None = None,
) -> dict[str, object]:
    capped_search_rows = search_rows[:max_rows]
    return {
        "ok": True,
        "generated_at": datetime.now().strftime(TIME_FORMAT),
        "settings": settings,
        "candidate_count": candidate_count,
        "result_count": result_count,
        "search_row_count": len(search_rows),
        "baseline_row_count": len(baseline_rows),
        "top_weight_candidate": capped_search_rows[0] if capped_search_rows else {},
        "search_rows": capped_search_rows,
        "baseline_rows": baseline_rows,
        "walk_forward": walk_forward or {"enabled": False},
        "regime_search": regime_search or {"enabled": False},
    }


def write_weight_search_json(path: str | Path, payload: dict[str, object]) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_weight_search_markdown(path: str | Path, payload: dict[str, object]) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(format_weight_search_markdown(payload), encoding="utf-8")


def format_weight_search_markdown(payload: dict[str, object]) -> str:
    settings = payload.get("settings") if isinstance(payload.get("settings"), dict) else {}
    top = payload.get("top_weight_candidate") if isinstance(payload.get("top_weight_candidate"), dict) else {}
    walk_forward = payload.get("walk_forward") if isinstance(payload.get("walk_forward"), dict) else {}
    walk_aggregate = walk_forward.get("aggregate") if isinstance(walk_forward.get("aggregate"), dict) else {}
    regime_search_payload = payload.get("regime_search") if isinstance(payload.get("regime_search"), dict) else {}
    lines = [
        "# Score Weight Search",
        "",
        f"- Generated at: {payload.get('generated_at', '')}",
        f"- Side: {settings.get('side', '')}",
        f"- RR: {settings.get('rr', '')}",
        f"- Candidate/result count: {payload.get('candidate_count', 0)} / {payload.get('result_count', 0)}",
        f"- Search rows: {payload.get('search_row_count', 0)}",
        f"- Min count: {settings.get('min_count', '')}",
        "",
        "## Top Candidate",
        "",
    ]
    if top:
        for key in (
            "side",
            "threshold",
            "weights",
            "count",
            "wins",
            "losses",
            "timeouts",
            "win_rate",
            "avg_r",
            "pf",
            "total_r",
            "max_losing_streak",
            "max_drawdown_r",
        ):
            lines.append(f"- {key}: {top.get(key, '')}")
    else:
        lines.append("- None.")
    lines.extend(
        [
            "",
            "## Walk Forward",
            "",
        ]
    )
    if walk_forward.get("enabled") is True:
        for key in (
            "status",
            "folds",
            "folds_with_weight_trades",
            "required_folds_with_weight_trades",
            "missing_folds_with_weight_trades",
            "total_test_weight_count",
            "required_test_weight_count",
            "missing_test_weight_count",
            "min_test_weight_count",
            "min_test_weight_fold",
            "mean_test_weight_avg_r",
            "mean_test_baseline_avg_r",
            "mean_test_weight_pf",
            "mean_test_baseline_pf",
            "total_test_weight_r",
            "total_test_baseline_r",
            "delta_total_r",
            "recommendation",
        ):
            lines.append(f"- {key}: {walk_aggregate.get(key, '')}")
    else:
        lines.append("- Disabled.")
    lines.extend(["", "## Top Search Rows", ""])
    append_markdown_table(
        lines,
        payload.get("search_rows"),
        ("threshold", "weights", "count", "avg_r", "pf", "total_r", "max_drawdown_r", "max_losing_streak"),
        limit=10,
    )
    lines.extend(["", "## Baseline Thresholds", ""])
    append_markdown_table(
        lines,
        payload.get("baseline_rows"),
        ("threshold", "count", "avg_r", "pf", "total_r", "max_drawdown_r"),
        limit=10,
    )
    if regime_search_payload.get("enabled"):
        lines.extend(["", "## Regime Search", ""])
        best = (
            regime_search_payload.get("best_regime_candidate")
            if isinstance(regime_search_payload.get("best_regime_candidate"), dict)
            else {}
        )
        if best:
            lines.append(
                "- Best regime candidate: "
                f"{best.get('dimension', '')}={best.get('group', '')}, "
                f"threshold={best.get('threshold', '')}, count={best.get('count', '')}, "
                f"avg_r={best.get('avg_r', '')}, pf={best.get('pf', '')}, "
                f"wf={regime_walk_forward_status(best.get('walk_forward'))}"
            )
        lines.append(f"- Rows: {regime_search_payload.get('row_count', 0)}")
        lines.append(f"- Skipped groups: {regime_search_payload.get('skipped_group_count', 0)}")
        lines.extend(["", "### Regime Candidates", ""])
        append_regime_markdown_table(lines, regime_search_payload.get("rows"), limit=12)
        lines.extend(["", "### Skipped Regime Groups", ""])
        append_markdown_table(
            lines,
            regime_search_payload.get("skipped_groups"),
            ("dimension", "group", "count", "status"),
            limit=20,
        )
    return "\n".join(lines).rstrip() + "\n"


def append_regime_markdown_table(lines: list[str], rows: object, *, limit: int) -> None:
    table_rows = []
    for row in rows if isinstance(rows, list) else []:
        if not isinstance(row, dict):
            continue
        walk = row.get("walk_forward") if isinstance(row.get("walk_forward"), dict) else {}
        aggregate = walk.get("aggregate") if isinstance(walk.get("aggregate"), dict) else {}
        deltas = row.get("deltas") if isinstance(row.get("deltas"), dict) else {}
        table_rows.append(
            {
                "dimension": row.get("dimension", ""),
                "group": row.get("group", ""),
                "threshold": row.get("threshold", ""),
                "weights": row.get("weights", ""),
                "count": row.get("count", ""),
                "avg_r": row.get("avg_r", ""),
                "pf": row.get("pf", ""),
                "delta_avg_r": deltas.get("avg_r", ""),
                "delta_pf": deltas.get("pf", ""),
                "wf_status": aggregate.get("status", ""),
                "wf_weight_count": aggregate.get("total_test_weight_count", ""),
                "wf_missing_count": aggregate.get("missing_test_weight_count", ""),
                "wf_delta_total_r": aggregate.get("delta_total_r", ""),
            }
        )
    append_markdown_table(
        lines,
        table_rows,
        (
            "dimension",
            "group",
            "threshold",
            "weights",
            "count",
            "avg_r",
            "pf",
            "delta_avg_r",
            "delta_pf",
            "wf_status",
            "wf_weight_count",
            "wf_missing_count",
            "wf_delta_total_r",
        ),
        limit=limit,
    )


def append_markdown_table(lines: list[str], rows: object, headers: tuple[str, ...], *, limit: int) -> None:
    if not isinstance(rows, list) or not rows:
        lines.append("- None.")
        return
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("|" + "|".join("---" for _ in headers) + "|")
    for row in rows[:limit]:
        if not isinstance(row, dict):
            continue
        values = [markdown_cell(row.get(header, "")) for header in headers]
        lines.append("| " + " | ".join(values) + " |")


def markdown_cell(value: object) -> str:
    if isinstance(value, float):
        value = round(value, 4)
    text = str(value if value is not None else "")
    return text.replace("|", "\\|").replace("\n", " ")


def regime_walk_forward_status(value: object) -> str:
    if not isinstance(value, dict):
        return ""
    aggregate = value.get("aggregate") if isinstance(value.get("aggregate"), dict) else {}
    return str(aggregate.get("status", ""))


def weight_search_walk_forward(
    candidates: list[Candidate],
    results: list[BacktestResult],
    *,
    side: str,
    thresholds=DEFAULT_THRESHOLDS,
    min_count: int,
    folds: int,
    train_window: int,
    test_window: int,
    purge_records: int = 0,
    embargo_records: int = 0,
    embargo_minutes: int = 0,
) -> dict[str, object]:
    pairs = candidate_result_pairs(candidates, results, side=side)
    windows = pair_walk_forward_windows(
        pairs,
        folds=folds,
        train_window=train_window,
        test_window=test_window,
        purge_records=purge_records,
        embargo_records=embargo_records,
        embargo_minutes=embargo_minutes,
    )
    fold_rows: list[dict[str, object]] = []
    for fold_index, (train, test) in enumerate(windows, start=1):
        train_candidates = [candidate for candidate, _ in train]
        train_results = [result for _, result in train]
        train_rows = component_weight_search(
            train_candidates,
            train_results,
            side=side,
            thresholds=thresholds,
            min_count=min_count,
        )
        baseline_rows = threshold_diagnostics(train_candidates, train_results, thresholds=thresholds)
        baseline = select_baseline_row(baseline_rows, min_count=min_count)
        if train_rows:
            selected = train_rows[0]
            train_weight_metrics = evaluate_weight_row(train, selected)
            test_weight_metrics = evaluate_weight_row(test, selected)
        else:
            selected = {}
            train_weight_metrics = metrics_for_pairs([])
            test_weight_metrics = metrics_for_pairs([])
        test_baseline_metrics = evaluate_baseline_row(test, baseline)
        fold_rows.append(
            {
                "fold": fold_index,
                "train_start": train[0][0].time_text if train else "",
                "train_end": train[-1][0].time_text if train else "",
                "test_start": test[0][0].time_text if test else "",
                "test_end": test[-1][0].time_text if test else "",
                "selected_weights": selected.get("weights", ""),
                "selected_threshold": selected.get("threshold", ""),
                "baseline_threshold": baseline.get("threshold", ""),
                **prefixed_metrics("train_weight_", train_weight_metrics),
                **prefixed_metrics("test_weight_", test_weight_metrics),
                **prefixed_metrics("test_baseline_", test_baseline_metrics),
                "delta_test_count": int(test_weight_metrics.get("count", 0) or 0)
                - int(test_baseline_metrics.get("count", 0) or 0),
                "delta_test_avg_r": round(float(test_weight_metrics.get("avg_r", 0.0) or 0.0) - float(test_baseline_metrics.get("avg_r", 0.0) or 0.0), 4),
                "delta_test_pf": round(float(test_weight_metrics.get("pf", 0.0) or 0.0) - float(test_baseline_metrics.get("pf", 0.0) or 0.0), 4),
                "delta_test_total_r": round(float(test_weight_metrics.get("total_r", 0.0) or 0.0) - float(test_baseline_metrics.get("total_r", 0.0) or 0.0), 4),
                "delta_test_max_drawdown_r": round(
                    float(test_weight_metrics.get("max_drawdown_r", 0.0) or 0.0)
                    - float(test_baseline_metrics.get("max_drawdown_r", 0.0) or 0.0),
                    4,
                ),
            }
        )
    aggregate = aggregate_weight_walk_forward_rows(fold_rows, min_count=min_count)
    return {
        "enabled": True,
        "settings": {
            "folds": folds,
            "train_window": train_window,
            "test_window": test_window,
            "purge_records": purge_records,
            "embargo_records": embargo_records,
            "embargo_minutes": embargo_minutes,
            "min_count": min_count,
        },
        "fold_rows": fold_rows,
        "aggregate": aggregate,
    }


def candidate_result_pairs(
    candidates: list[Candidate],
    results: list[BacktestResult],
    *,
    side: str,
) -> list[tuple[Candidate, BacktestResult]]:
    result_by_id = {result.candidate_id: result for result in results}
    pairs = [
        (candidate, result_by_id[candidate.candidate_id])
        for candidate in candidates
        if candidate.candidate_id in result_by_id and (side == "both" or candidate.side == side)
    ]
    return sorted(pairs, key=lambda pair: pair[0].time)


def pair_walk_forward_windows(
    pairs: list[tuple[Candidate, BacktestResult]],
    *,
    folds: int,
    train_window: int,
    test_window: int,
    purge_records: int = 0,
    embargo_records: int = 0,
    embargo_minutes: int = 0,
) -> list[tuple[list[tuple[Candidate, BacktestResult]], list[tuple[Candidate, BacktestResult]]]]:
    if len(pairs) < 4:
        return []
    purge_records = max(0, purge_records)
    embargo_records = max(0, embargo_records)
    embargo_minutes = max(0, embargo_minutes)
    max_train_window = len(pairs) - embargo_records - 1
    if max_train_window < 2:
        return []
    train_window = max(2, min(train_window, max_train_window))
    test_window = max(1, min(test_window, len(pairs) - train_window - embargo_records))
    possible_starts = list(range(0, len(pairs) - train_window - embargo_records - test_window + 1))
    if not possible_starts:
        train, test = apply_pair_purge_embargo(
            pairs[:train_window],
            pairs[train_window + embargo_records :],
            purge_records=purge_records,
            embargo_minutes=embargo_minutes,
            split_train_end=pairs[train_window - 1],
            split_test_start=pairs[train_window] if train_window < len(pairs) else None,
        )
        return [(train, test)] if train and test else []
    starts = [possible_starts[-1]] if folds <= 1 else [
        possible_starts[round(index * (len(possible_starts) - 1) / (folds - 1))]
        for index in range(folds)
    ]
    windows: list[tuple[list[tuple[Candidate, BacktestResult]], list[tuple[Candidate, BacktestResult]]]] = []
    seen: set[int] = set()
    for start in starts:
        if start in seen:
            continue
        seen.add(start)
        train_end = start + train_window
        test_start = train_end + embargo_records
        train = pairs[start:train_end]
        test = pairs[test_start : test_start + test_window]
        train, test = apply_pair_purge_embargo(
            train,
            test,
            purge_records=purge_records,
            embargo_minutes=embargo_minutes,
            split_train_end=pairs[train_end - 1],
            split_test_start=pairs[train_end] if train_end < len(pairs) else None,
        )
        if train and test:
            windows.append((train, test))
    return windows


def apply_pair_purge_embargo(
    train: list[tuple[Candidate, BacktestResult]],
    test: list[tuple[Candidate, BacktestResult]],
    *,
    purge_records: int,
    embargo_minutes: int,
    split_train_end: tuple[Candidate, BacktestResult] | None,
    split_test_start: tuple[Candidate, BacktestResult] | None,
) -> tuple[list[tuple[Candidate, BacktestResult]], list[tuple[Candidate, BacktestResult]]]:
    if purge_records > 0:
        train = train[:-purge_records] if len(train) > purge_records else []
    if embargo_minutes <= 0:
        return train, test
    if split_test_start is not None:
        test_start_time = split_test_start[0].time
        train = [
            pair
            for pair in train
            if (test_start_time - pair[0].time).total_seconds() > embargo_minutes * 60
        ]
    if split_train_end is not None:
        train_end_time = split_train_end[0].time
        test = [pair for pair in test if (pair[0].time - train_end_time).total_seconds() > embargo_minutes * 60]
    return train, test


def select_baseline_row(rows: list[dict[str, object]], *, min_count: int) -> dict[str, object]:
    sufficient = [row for row in rows if int(row.get("count", 0) or 0) >= min_count]
    if not sufficient:
        sufficient = [row for row in rows if int(row.get("count", 0) or 0) > 0]
    if not sufficient:
        return {}
    return max(
        sufficient,
        key=lambda row: (
            float(row.get("avg_r", 0.0) or 0.0),
            float(row.get("pf", 0.0) or 0.0),
            float(row.get("total_r", 0.0) or 0.0),
        ),
    )


def weights_from_row(row: dict[str, object]) -> ComponentWeights:
    return ComponentWeights(
        trend=float(row.get("trend_w", 1.0) or 1.0),
        structure=float(row.get("structure_w", 1.0) or 1.0),
        entry=float(row.get("entry_w", 1.0) or 1.0),
        risk=float(row.get("risk_w", 1.0) or 1.0),
        cost=float(row.get("cost_w", 1.0) or 1.0),
        chop=float(row.get("chop_w", 1.0) or 1.0),
    )


def evaluate_weight_row(pairs: list[tuple[Candidate, BacktestResult]], row: dict[str, object]) -> dict[str, object]:
    if not row:
        return metrics_for_pairs([])
    side = str(row.get("side") or "both")
    threshold = float(row.get("threshold", 0.0) or 0.0)
    weights = weights_from_row(row)
    selected = [
        (candidate, result)
        for candidate, result in pairs
        if (side == "both" or candidate.side == side) and weighted_score(candidate, weights) >= threshold
    ]
    return metrics_for_pairs(selected)


def evaluate_baseline_row(pairs: list[tuple[Candidate, BacktestResult]], row: dict[str, object]) -> dict[str, object]:
    if not row:
        return metrics_for_pairs([])
    threshold = float(row.get("threshold", 0.0) or 0.0)
    return metrics_for_pairs([(candidate, result) for candidate, result in pairs if candidate.score >= threshold])


def prefixed_metrics(prefix: str, row: dict[str, object]) -> dict[str, object]:
    return {f"{prefix}{key}": value for key, value in row.items()}


def aggregate_weight_walk_forward_rows(rows: list[dict[str, object]], *, min_count: int) -> dict[str, object]:
    actual = [row for row in rows if row.get("fold") != "aggregate"]
    if not actual:
        return {
            "status": "no_walk_forward_windows",
            "folds": 0,
            "folds_with_weight_trades": 0,
            "required_folds_with_weight_trades": 1,
            "total_test_weight_count": 0,
            "required_test_weight_count": min_count,
            "missing_test_weight_count": min_count,
            "missing_folds_with_weight_trades": 1,
            "folds_without_weight_trades": 0,
            "recommendation": "increase sample size before using weight search as adoption evidence",
        }

    def values(key: str, *, count_key: str) -> list[float]:
        return [float(row.get(key, 0.0) or 0.0) for row in actual if int(row.get(count_key, 0) or 0) > 0]

    test_weight_counts = [int(row.get("test_weight_count", 0) or 0) for row in actual]
    test_baseline_counts = [int(row.get("test_baseline_count", 0) or 0) for row in actual]
    weight_avg_rs = values("test_weight_avg_r", count_key="test_weight_count")
    baseline_avg_rs = values("test_baseline_avg_r", count_key="test_baseline_count")
    weight_pfs = values("test_weight_pf", count_key="test_weight_count")
    baseline_pfs = values("test_baseline_pf", count_key="test_baseline_count")
    total_weight_r = round(sum(float(row.get("test_weight_total_r", 0.0) or 0.0) for row in actual), 4)
    total_baseline_r = round(sum(float(row.get("test_baseline_total_r", 0.0) or 0.0) for row in actual), 4)
    mean_weight_avg_r = round(mean(weight_avg_rs), 4) if weight_avg_rs else 0.0
    mean_baseline_avg_r = round(mean(baseline_avg_rs), 4) if baseline_avg_rs else 0.0
    mean_weight_pf = round(mean(weight_pfs), 4) if weight_pfs else 0.0
    mean_baseline_pf = round(mean(baseline_pfs), 4) if baseline_pfs else 0.0
    folds_with_weight = sum(1 for count in test_weight_counts if count > 0)
    required_folds = max(1, len(actual) // 2)
    total_weight_count = sum(test_weight_counts)
    delta_total_r = round(total_weight_r - total_baseline_r, 4)
    min_weight_count = min(test_weight_counts) if test_weight_counts else 0
    min_weight_fold = ""
    if actual and test_weight_counts:
        min_index = test_weight_counts.index(min_weight_count)
        min_weight_fold = actual[min_index].get("fold", "")
    folds_without_weight = sum(1 for count in test_weight_counts if count == 0)
    missing_test_weight_count = max(0, min_count - total_weight_count)
    missing_folds_with_weight = max(0, required_folds - folds_with_weight)
    improved = (
        mean_weight_avg_r > mean_baseline_avg_r
        and mean_weight_pf >= mean_baseline_pf
        and mean_weight_avg_r > 0
        and mean_weight_pf >= 1.0
        and total_weight_r > 0
        and delta_total_r > 0
    )
    enough_samples = total_weight_count >= min_count and folds_with_weight >= required_folds
    if enough_samples and improved:
        status = "walk_forward_candidate_passed"
        recommendation = "candidate can proceed to MT5 optimization/yearly validation; still not adopted"
    elif not enough_samples:
        status = "walk_forward_sample_shortage"
        recommendation = "collect more samples before treating this fitted weighting as robust"
    else:
        status = "walk_forward_candidate_failed"
        recommendation = "do not adopt this weighting; fold test did not improve over baseline"
    return {
        "status": status,
        "folds": len(actual),
        "folds_with_weight_trades": folds_with_weight,
        "required_folds_with_weight_trades": required_folds,
        "total_test_weight_count": total_weight_count,
        "required_test_weight_count": min_count,
        "missing_test_weight_count": missing_test_weight_count,
        "total_test_baseline_count": sum(test_baseline_counts),
        "missing_folds_with_weight_trades": missing_folds_with_weight,
        "folds_without_weight_trades": folds_without_weight,
        "min_test_weight_count": min_weight_count,
        "min_test_weight_fold": min_weight_fold,
        "mean_test_weight_avg_r": mean_weight_avg_r,
        "mean_test_baseline_avg_r": mean_baseline_avg_r,
        "mean_test_weight_pf": mean_weight_pf,
        "mean_test_baseline_pf": mean_baseline_pf,
        "median_test_weight_avg_r": round(median(weight_avg_rs), 4) if weight_avg_rs else 0.0,
        "median_test_baseline_avg_r": round(median(baseline_avg_rs), 4) if baseline_avg_rs else 0.0,
        "median_test_weight_pf": round(median(weight_pfs), 4) if weight_pfs else 0.0,
        "median_test_baseline_pf": round(median(baseline_pfs), 4) if baseline_pfs else 0.0,
        "total_test_weight_r": total_weight_r,
        "total_test_baseline_r": total_baseline_r,
        "delta_mean_avg_r": round(mean_weight_avg_r - mean_baseline_avg_r, 4),
        "delta_mean_pf": round(mean_weight_pf - mean_baseline_pf, 4),
        "delta_total_r": delta_total_r,
        "min_count": min_count,
        "recommendation": recommendation,
    }


def regime_weight_search(
    candidates: list[Candidate],
    results: list[BacktestResult],
    *,
    side: str,
    regimes: list[str],
    thresholds=DEFAULT_THRESHOLDS,
    min_count: int,
    top_per_group: int = 1,
    walk_forward: bool = False,
    wf_folds: int = 4,
    wf_train_window: int = 240,
    wf_test_window: int = 60,
    wf_purge_records: int = 0,
    wf_embargo_records: int = 0,
    wf_embargo_minutes: int = 0,
) -> dict[str, object]:
    pairs = candidate_result_pairs(candidates, results, side=side)
    rows: list[dict[str, object]] = []
    skipped_groups: list[dict[str, object]] = []
    for dimension in regimes:
        groups: dict[str, list[tuple[Candidate, BacktestResult]]] = {}
        for candidate, result in pairs:
            group = candidate_regime_value(candidate, dimension)
            groups.setdefault(group, []).append((candidate, result))
        for group, group_pairs in sorted(groups.items(), key=lambda item: item[0]):
            if len(group_pairs) < min_count:
                skipped_groups.append(
                    {
                        "dimension": dimension,
                        "group": group,
                        "count": len(group_pairs),
                        "status": "sample_shortage",
                    }
                )
                continue
            group_candidates = [candidate for candidate, _ in group_pairs]
            group_results = [result for _, result in group_pairs]
            search_rows = component_weight_search(
                group_candidates,
                group_results,
                side=side,
                thresholds=thresholds,
                min_count=min_count,
            )
            baseline_rows = threshold_diagnostics(group_candidates, group_results, thresholds=thresholds)
            baseline = select_baseline_row(baseline_rows, min_count=min_count)
            if not search_rows:
                skipped_groups.append(
                    {
                        "dimension": dimension,
                        "group": group,
                        "count": len(group_pairs),
                        "status": "no_weight_candidate",
                    }
                )
                continue
            for rank, row in enumerate(search_rows[: max(1, top_per_group)], start=1):
                walk = (
                    weight_search_walk_forward(
                        group_candidates,
                        group_results,
                        side=side,
                        thresholds=thresholds,
                        min_count=min_count,
                        folds=wf_folds,
                        train_window=wf_train_window,
                        test_window=wf_test_window,
                        purge_records=wf_purge_records,
                        embargo_records=wf_embargo_records,
                        embargo_minutes=wf_embargo_minutes,
                    )
                    if walk_forward
                    else {"enabled": False}
                )
                rows.append(
                    {
                        "dimension": dimension,
                        "group": group,
                        "rank": rank,
                        **row,
                        "baseline": compact_baseline_row(baseline),
                        "deltas": score_weight_row_deltas(row, baseline),
                        "walk_forward": walk,
                    }
                )
    best = sorted(
        rows,
        key=lambda row: (
            regime_walk_forward_pass_rank(row.get("walk_forward")),
            float(row.get("avg_r", 0.0) or 0.0),
            float(row.get("pf", 0.0) or 0.0),
            float(row.get("total_r", 0.0) or 0.0),
        ),
        reverse=True,
    )
    return {
        "enabled": bool(regimes),
        "settings": {
            "side": side,
            "regimes": regimes,
            "min_count": min_count,
            "top_per_group": top_per_group,
            "walk_forward": walk_forward,
            "wf_folds": wf_folds,
            "wf_train_window": wf_train_window,
            "wf_test_window": wf_test_window,
            "wf_purge_records": wf_purge_records,
            "wf_embargo_records": wf_embargo_records,
            "wf_embargo_minutes": wf_embargo_minutes,
        },
        "row_count": len(rows),
        "skipped_group_count": len(skipped_groups),
        "best_regime_candidate": compact_regime_weight_row(best[0]) if best else {},
        "rows": [compact_regime_weight_row(row) for row in rows[:200]],
        "skipped_groups": skipped_groups[:200],
    }


def candidate_regime_value(candidate: Candidate, dimension: str) -> str:
    if dimension == "entry_hour":
        hour = candidate.time.hour
        return f"{hour:02d}:00-{(hour + 1) % 24:02d}:00"
    if dimension == "m30_trend":
        return timeframe_trend(candidate, "M30")
    if dimension == "m15_trend":
        return timeframe_trend(candidate, "M15")
    if dimension == "m30_m15_trend":
        return f"M30 {timeframe_trend(candidate, 'M30')} M15 {timeframe_trend(candidate, 'M15')}"
    if dimension == "htf_alignment":
        alignment = candidate.features.get("htf_alignment_count")
        slope = candidate.features.get("htf_slope_count")
        if _is_number(alignment) and _is_number(slope):
            aligned = float(alignment) >= 3.0 and float(slope) >= 2.0
            return "aligned" if aligned else "mixed_or_against"
        return "unknown"
    return "unknown"


def timeframe_trend(candidate: Candidate, timeframe: str) -> str:
    fast = _as_float(candidate.features.get(f"{timeframe}_ema_fast"))
    slow = _as_float(candidate.features.get(f"{timeframe}_ema_slow"))
    if fast is None or slow is None:
        return "unknown"
    if fast > slow:
        return "up"
    if fast < slow:
        return "down"
    return "flat"


def score_weight_row_deltas(row: dict[str, object], baseline: dict[str, object]) -> dict[str, object]:
    if not baseline:
        return {}
    return {
        "count": int(row.get("count", 0) or 0) - int(baseline.get("count", 0) or 0),
        "avg_r": round(float(row.get("avg_r", 0.0) or 0.0) - float(baseline.get("avg_r", 0.0) or 0.0), 4),
        "pf": round(float(row.get("pf", 0.0) or 0.0) - float(baseline.get("pf", 0.0) or 0.0), 4),
        "total_r": round(float(row.get("total_r", 0.0) or 0.0) - float(baseline.get("total_r", 0.0) or 0.0), 4),
        "max_drawdown_r": round(
            float(row.get("max_drawdown_r", 0.0) or 0.0) - float(baseline.get("max_drawdown_r", 0.0) or 0.0),
            4,
        ),
    }


def compact_baseline_row(row: dict[str, object]) -> dict[str, object]:
    return {
        "threshold": row.get("threshold"),
        "count": row.get("count"),
        "avg_r": row.get("avg_r"),
        "pf": row.get("pf"),
        "total_r": row.get("total_r"),
        "max_drawdown_r": row.get("max_drawdown_r"),
    } if row else {}


def compact_regime_weight_row(row: dict[str, object]) -> dict[str, object]:
    walk = row.get("walk_forward")
    return {
        "dimension": row.get("dimension"),
        "group": row.get("group"),
        "rank": row.get("rank"),
        "side": row.get("side"),
        "threshold": row.get("threshold"),
        "weights": row.get("weights"),
        "count": row.get("count"),
        "win_rate": row.get("win_rate"),
        "avg_r": row.get("avg_r"),
        "pf": row.get("pf"),
        "total_r": row.get("total_r"),
        "max_losing_streak": row.get("max_losing_streak"),
        "max_drawdown_r": row.get("max_drawdown_r"),
        "trend_w": row.get("trend_w"),
        "structure_w": row.get("structure_w"),
        "entry_w": row.get("entry_w"),
        "risk_w": row.get("risk_w"),
        "cost_w": row.get("cost_w"),
        "chop_w": row.get("chop_w"),
        "baseline": row.get("baseline") if isinstance(row.get("baseline"), dict) else {},
        "deltas": row.get("deltas") if isinstance(row.get("deltas"), dict) else {},
        "walk_forward": compact_regime_walk_forward(walk),
    }


def compact_regime_walk_forward(value: object) -> dict[str, object]:
    if not isinstance(value, dict):
        return {}
    aggregate = value.get("aggregate") if isinstance(value.get("aggregate"), dict) else {}
    if value.get("enabled") is not True:
        return {"enabled": False}
    return {
        "enabled": True,
        "aggregate": {
            key: aggregate.get(key)
            for key in (
                "status",
                "folds",
                "total_test_weight_count",
                "required_test_weight_count",
                "missing_test_weight_count",
                "total_test_baseline_count",
                "folds_with_weight_trades",
                "required_folds_with_weight_trades",
                "missing_folds_with_weight_trades",
                "folds_without_weight_trades",
                "min_test_weight_count",
                "min_test_weight_fold",
                "mean_test_weight_avg_r",
                "mean_test_baseline_avg_r",
                "mean_test_weight_pf",
                "mean_test_baseline_pf",
                "total_test_weight_r",
                "total_test_baseline_r",
                "delta_mean_avg_r",
                "delta_mean_pf",
                "delta_total_r",
                "min_count",
                "recommendation",
            )
        },
    }


def regime_walk_forward_pass_rank(value: object) -> int:
    if not isinstance(value, dict) or value.get("enabled") is not True:
        return 0
    aggregate = value.get("aggregate") if isinstance(value.get("aggregate"), dict) else {}
    status = aggregate.get("status")
    if status == "walk_forward_candidate_passed":
        return 3
    if status == "walk_forward_candidate_failed":
        return 1
    return 0


def flatten_regime_rows(regime_search: dict[str, object]) -> list[dict[str, object]]:
    rows = regime_search.get("rows") if isinstance(regime_search, dict) else []
    flattened: list[dict[str, object]] = []
    iterable = rows if isinstance(rows, list) else []
    for row in iterable:
        if not isinstance(row, dict):
            continue
        walk = row.get("walk_forward") if isinstance(row.get("walk_forward"), dict) else {}
        aggregate = walk.get("aggregate") if isinstance(walk.get("aggregate"), dict) else {}
        baseline = row.get("baseline") if isinstance(row.get("baseline"), dict) else {}
        deltas = row.get("deltas") if isinstance(row.get("deltas"), dict) else {}
        flattened.append(
            {
                **{key: value for key, value in row.items() if key not in {"walk_forward", "baseline", "deltas"}},
                "baseline_threshold": baseline.get("threshold"),
                "baseline_count": baseline.get("count"),
                "baseline_avg_r": baseline.get("avg_r"),
                "baseline_pf": baseline.get("pf"),
                "baseline_total_r": baseline.get("total_r"),
                "delta_count": deltas.get("count"),
                "delta_avg_r": deltas.get("avg_r"),
                "delta_pf": deltas.get("pf"),
                "delta_total_r": deltas.get("total_r"),
                "wf_status": aggregate.get("status"),
                "wf_weight_count": aggregate.get("total_test_weight_count"),
                "wf_required_weight_count": aggregate.get("required_test_weight_count"),
                "wf_missing_weight_count": aggregate.get("missing_test_weight_count"),
                "wf_baseline_count": aggregate.get("total_test_baseline_count"),
                "wf_folds_with_weight": aggregate.get("folds_with_weight_trades"),
                "wf_required_folds_with_weight": aggregate.get("required_folds_with_weight_trades"),
                "wf_missing_folds_with_weight": aggregate.get("missing_folds_with_weight_trades"),
                "wf_min_weight_count": aggregate.get("min_test_weight_count"),
                "wf_min_weight_fold": aggregate.get("min_test_weight_fold"),
                "wf_mean_avg_r": aggregate.get("mean_test_weight_avg_r"),
                "wf_baseline_avg_r": aggregate.get("mean_test_baseline_avg_r"),
                "wf_mean_pf": aggregate.get("mean_test_weight_pf"),
                "wf_baseline_pf": aggregate.get("mean_test_baseline_pf"),
                "wf_delta_total_r": aggregate.get("delta_total_r"),
            }
        )
    return flattened


def parse_regime_list(value: str) -> list[str]:
    allowed = {"entry_hour", "m30_trend", "m15_trend", "m30_m15_trend", "htf_alignment"}
    regimes = []
    for item in str(value or "").split(","):
        name = item.strip()
        if not name:
            continue
        if name not in allowed:
            raise argparse.ArgumentTypeError(f"unsupported regime: {name}")
        regimes.append(name)
    return regimes


def _as_float(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _is_number(value: object) -> bool:
    return _as_float(value) is not None


def _append_rows(ws, rows: list[dict[str, object]]) -> None:
    if not rows:
        return
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])


def _format_sheet(ws) -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
    ws.freeze_panes = "A2"
    for col in range(1, ws.max_column + 1):
        width = max(len(str(ws.cell(row, col).value or "")) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), 36)
    if ws.max_row > 1 and ws.max_column > 1:
        safe_title = "".join(ch if ch.isascii() and ch.isalnum() else "_" for ch in ws.title)
        table = Table(displayName=f"Table_{safe_title}", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)


def _write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Search score component weights for swing candidates.")
    parser.add_argument("--history", default="runtime/latest_history_168h.json")
    parser.add_argument("--output", default="reports/score_weight_search.xlsx")
    parser.add_argument("--output-json", default="")
    parser.add_argument("--output-md", default="")
    parser.add_argument("--rr", type=float, default=5.0)
    parser.add_argument("--side", choices=("buy", "sell", "both"), default="both")
    parser.add_argument("--min-count", type=int, default=30)
    parser.add_argument("--max-rows", type=int, default=200)
    parser.add_argument("--max-hold-minutes", type=int, default=60)
    parser.add_argument("--swing-left", type=int, default=3)
    parser.add_argument("--swing-right", type=int, default=3)
    parser.add_argument("--min-atr-distance", type=float, default=0.5)
    parser.add_argument("--max-risk-atr", type=float, default=3.0)
    parser.add_argument("--score-profile", choices=("side", "balanced", "buy", "sell"), default="side")
    parser.add_argument("--include-blackout-times", action="store_true", help="Include rollover/news no-entry windows in candidate generation.")
    parser.add_argument("--calendar", default="runtime/economic_calendar.json", help="Optional economic calendar JSON/CSV in MT5 server time.")
    parser.add_argument("--calendar-input-utc-offset", type=float, default=None, help="UTC offset of naive calendar times, e.g. 9 for JST. Omit when calendar is already MT5 server time.")
    parser.add_argument("--calendar-server-utc-offset", type=float, default=None, help="MT5 server UTC offset used when converting calendar times.")
    parser.add_argument("--news-before-minutes", type=int, default=10)
    parser.add_argument("--news-after-minutes", type=int, default=10)
    parser.add_argument("--news-min-impact", default="high", choices=("low", "medium", "high"))
    parser.add_argument("--news-currencies", default="USD,XAU,ALL")
    parser.add_argument("--walk-forward", action="store_true", help="Run chronological train/test validation for fitted score weights.")
    parser.add_argument("--wf-folds", type=int, default=4)
    parser.add_argument("--wf-train-window", type=int, default=240)
    parser.add_argument("--wf-test-window", type=int, default=60)
    parser.add_argument("--wf-purge-records", type=int, default=0)
    parser.add_argument("--wf-embargo-records", type=int, default=0)
    parser.add_argument("--wf-embargo-minutes", type=int, default=0)
    parser.add_argument(
        "--regime-search",
        type=parse_regime_list,
        default=[],
        help="Comma-separated regimes for separate weight search: entry_hour,m30_trend,m15_trend,m30_m15_trend,htf_alignment.",
    )
    parser.add_argument("--regime-min-count", type=int, default=0)
    parser.add_argument("--regime-top-per-group", type=int, default=1)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    history = load_history(args.history)
    calendar_events = load_economic_calendar(
        args.calendar,
        input_utc_offset_hours=args.calendar_input_utc_offset,
        server_utc_offset_hours=args.calendar_server_utc_offset,
    )
    candidates = generate_candidates(
        history,
        risk_reward=args.rr,
        swing_left=args.swing_left,
        swing_right=args.swing_right,
        min_atr_distance=args.min_atr_distance,
        max_risk_atr=args.max_risk_atr,
        min_score=None,
        score_profile=args.score_profile,
        exclude_blackout_times=not args.include_blackout_times,
        blackout_events=calendar_events,
        news_before_minutes=args.news_before_minutes,
        news_after_minutes=args.news_after_minutes,
        news_min_impact=args.news_min_impact,
        news_currencies=parse_currencies(args.news_currencies),
    )
    if args.side != "both":
        candidates = [candidate for candidate in candidates if candidate.side == args.side]
    spread_price = history.spread_points * history.point
    results = run_backtest(candidates, history.bars("M1"), max_hold_minutes=args.max_hold_minutes, spread_price=spread_price)
    search_side = "both" if args.side == "both" else args.side
    search_rows = component_weight_search(candidates, results, side=search_side, thresholds=DEFAULT_THRESHOLDS, min_count=args.min_count)
    baseline_rows = threshold_diagnostics(candidates, results, thresholds=DEFAULT_THRESHOLDS)
    walk_forward = (
        weight_search_walk_forward(
            candidates,
            results,
            side=search_side,
            thresholds=DEFAULT_THRESHOLDS,
            min_count=args.min_count,
            folds=args.wf_folds,
            train_window=args.wf_train_window,
            test_window=args.wf_test_window,
            purge_records=args.wf_purge_records,
            embargo_records=args.wf_embargo_records,
            embargo_minutes=args.wf_embargo_minutes,
        )
        if args.walk_forward
        else {"enabled": False}
    )
    walk_forward_rows = None
    if isinstance(walk_forward, dict) and walk_forward.get("enabled") is True:
        fold_rows = walk_forward.get("fold_rows")
        aggregate = walk_forward.get("aggregate")
        walk_forward_rows = [*fold_rows, aggregate] if isinstance(fold_rows, list) and isinstance(aggregate, dict) else []
    regime_search = (
        regime_weight_search(
            candidates,
            results,
            side=search_side,
            regimes=args.regime_search,
            thresholds=DEFAULT_THRESHOLDS,
            min_count=args.regime_min_count or args.min_count,
            top_per_group=args.regime_top_per_group,
            walk_forward=args.walk_forward,
            wf_folds=args.wf_folds,
            wf_train_window=args.wf_train_window,
            wf_test_window=args.wf_test_window,
            wf_purge_records=args.wf_purge_records,
            wf_embargo_records=args.wf_embargo_records,
            wf_embargo_minutes=args.wf_embargo_minutes,
        )
        if args.regime_search
        else {"enabled": False}
    )
    regime_rows = flatten_regime_rows(regime_search) if isinstance(regime_search, dict) and regime_search.get("enabled") else None
    payload = build_weight_search_payload(
        settings={
            "history": args.history,
            "output": args.output,
            "rr": args.rr,
            "side": args.side,
            "search_side": search_side,
            "min_count": args.min_count,
            "max_rows": args.max_rows,
            "max_hold_minutes": args.max_hold_minutes,
            "swing_left": args.swing_left,
            "swing_right": args.swing_right,
            "min_atr_distance": args.min_atr_distance,
            "max_risk_atr": args.max_risk_atr,
            "score_profile": args.score_profile,
            "include_blackout_times": args.include_blackout_times,
            "calendar": args.calendar,
            "calendar_input_utc_offset": args.calendar_input_utc_offset,
            "calendar_server_utc_offset": args.calendar_server_utc_offset,
            "news_before_minutes": args.news_before_minutes,
            "news_after_minutes": args.news_after_minutes,
            "news_min_impact": args.news_min_impact,
            "news_currencies": args.news_currencies,
            "walk_forward": args.walk_forward,
            "wf_folds": args.wf_folds,
            "wf_train_window": args.wf_train_window,
            "wf_test_window": args.wf_test_window,
            "wf_purge_records": args.wf_purge_records,
            "wf_embargo_records": args.wf_embargo_records,
            "wf_embargo_minutes": args.wf_embargo_minutes,
            "regime_search": args.regime_search,
            "regime_min_count": args.regime_min_count or args.min_count,
            "regime_top_per_group": args.regime_top_per_group,
        },
        candidate_count=len(candidates),
        result_count=len(results),
        search_rows=search_rows,
        baseline_rows=baseline_rows,
        max_rows=args.max_rows,
        walk_forward=walk_forward,
        regime_search=regime_search,
    )
    write_weight_search_report(args.output, search_rows[: args.max_rows], baseline_rows, walk_forward_rows, regime_rows)
    if args.output_json:
        write_weight_search_json(args.output_json, payload)
    if args.output_md:
        write_weight_search_markdown(args.output_md, payload)
    for row in search_rows[:10]:
        print(row)
    print(f"wrote {args.output}")
    if args.output_json:
        print(f"wrote {args.output_json}")
    if args.output_md:
        print(f"wrote {args.output_md}")
    if isinstance(walk_forward, dict) and walk_forward.get("enabled") is True:
        print(f"walk_forward {walk_forward.get('aggregate', {})}")
    if isinstance(regime_search, dict) and regime_search.get("enabled") is True:
        print(f"regime_search rows={regime_search.get('row_count')} best={regime_search.get('best_regime_candidate')}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
