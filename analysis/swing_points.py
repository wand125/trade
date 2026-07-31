from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
import sys
from pathlib import Path

if __package__ is None or __package__ == "":
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from analysis.market_data import Bar, MarketHistory, load_history


@dataclass(frozen=True)
class SwingPoint:
    index: int
    kind: str
    time_text: str
    price: float
    atr: float
    strength_atr: float


def detect_swings(
    bars: list[Bar],
    atr_values: list[float | None] | None = None,
    *,
    left: int = 3,
    right: int = 3,
    min_atr_distance: float = 0.5,
) -> list[SwingPoint]:
    if left < 1 or right < 1:
        raise ValueError("left and right must be positive")
    if len(bars) < left + right + 1:
        return []

    atr_values = atr_values or [None] * len(bars)
    swings: list[SwingPoint] = []
    last_price_by_kind: dict[str, float] = {}

    for index in range(left, len(bars) - right):
        bar = bars[index]
        left_bars = bars[index - left : index]
        right_bars = bars[index + 1 : index + right + 1]
        atr = float(atr_values[index] or 0.0)
        min_distance = atr * min_atr_distance

        if bar.high > max(item.high for item in left_bars) and bar.high > max(item.high for item in right_bars):
            if _far_enough("high", bar.high, last_price_by_kind, min_distance):
                strength = _swing_strength(bar.high, [item.high for item in left_bars + right_bars], atr, high=True)
                swings.append(SwingPoint(index, "high", bar.time_text, bar.high, atr, strength))
                last_price_by_kind["high"] = bar.high

        if bar.low < min(item.low for item in left_bars) and bar.low < min(item.low for item in right_bars):
            if _far_enough("low", bar.low, last_price_by_kind, min_distance):
                strength = _swing_strength(bar.low, [item.low for item in left_bars + right_bars], atr, high=False)
                swings.append(SwingPoint(index, "low", bar.time_text, bar.low, atr, strength))
                last_price_by_kind["low"] = bar.low

    return sorted(swings, key=lambda item: item.index)


def swing_report_rows(
    history: MarketHistory,
    *,
    timeframes: tuple[str, ...] = ("M1", "M5"),
    left: int | None = None,
    right: int | None = None,
    min_atr_distance: float = 0.5,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for timeframe in timeframes:
        bars = history.bars(timeframe)
        if not bars:
            continue
        tf_left, tf_right = swing_window(timeframe, left=left, right=right)
        swings = detect_swings(
            bars,
            history.indicator(timeframe, "atr14"),
            left=tf_left,
            right=tf_right,
            min_atr_distance=min_atr_distance,
        )
        previous_by_kind: dict[str, SwingPoint] = {}
        for swing in swings:
            previous = previous_by_kind.get(swing.kind)
            confirm_index = min(len(bars) - 1, swing.index + tf_right)
            distance = abs(swing.price - previous.price) if previous is not None else 0.0
            denominator = swing.atr or (previous.atr if previous is not None else 0.0)
            rows.append(
                {
                    "symbol": history.symbol,
                    "timeframe": timeframe,
                    "index": swing.index,
                    "kind": swing.kind,
                    "swing_time": swing.time_text,
                    "confirmed_time": bars[confirm_index].time_text,
                    "confirmed_after_bars": tf_right,
                    "price": round(swing.price, 5),
                    "atr": round(swing.atr, 5),
                    "strength_atr": round(swing.strength_atr, 4),
                    "previous_same_kind_time": previous.time_text if previous is not None else "",
                    "previous_same_kind_price": round(previous.price, 5) if previous is not None else "",
                    "distance_from_previous_same_kind": round(distance, 5) if previous is not None else "",
                    "distance_from_previous_same_kind_atr": round(distance / denominator, 4)
                    if previous is not None and denominator > 0
                    else "",
                }
            )
            previous_by_kind[swing.kind] = swing
    return rows


def swing_summary_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    grouped: dict[tuple[str, str], list[dict[str, object]]] = {}
    for row in rows:
        key = (str(row.get("timeframe") or ""), str(row.get("kind") or ""))
        grouped.setdefault(key, []).append(row)

    summary: list[dict[str, object]] = []
    for (timeframe, kind), group_rows in sorted(grouped.items()):
        strengths = [float(row.get("strength_atr") or 0.0) for row in group_rows]
        distances = [
            float(row["distance_from_previous_same_kind_atr"])
            for row in group_rows
            if row.get("distance_from_previous_same_kind_atr") not in ("", None)
        ]
        summary.append(
            {
                "timeframe": timeframe,
                "kind": kind,
                "count": len(group_rows),
                "avg_strength_atr": round(sum(strengths) / len(strengths), 4) if strengths else 0.0,
                "avg_distance_from_previous_same_kind_atr": round(sum(distances) / len(distances), 4)
                if distances
                else 0.0,
                "first_swing_time": group_rows[0].get("swing_time", ""),
                "last_swing_time": group_rows[-1].get("swing_time", ""),
            }
        )
    return summary


def write_swing_report(path: str | Path, rows: list[dict[str, object]]) -> Path:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.suffix.lower() == ".xlsx":
        return _write_xlsx(output, rows)
    return _write_csv(output, rows)


def swing_window(timeframe: str, *, left: int | None = None, right: int | None = None) -> tuple[int, int]:
    if left is not None or right is not None:
        resolved_left = left if left is not None else right
        resolved_right = right if right is not None else left
        if resolved_left is None or resolved_right is None:
            raise ValueError("left/right resolution failed")
        return resolved_left, resolved_right
    if timeframe.upper() == "M1":
        return 3, 3
    if timeframe.upper() == "M5":
        return 2, 2
    return 2, 2


def _far_enough(kind: str, price: float, last_price_by_kind: dict[str, float], min_distance: float) -> bool:
    last = last_price_by_kind.get(kind)
    if last is None:
        return True
    return abs(price - last) >= min_distance


def _swing_strength(price: float, surrounding_prices: list[float], atr: float, *, high: bool) -> float:
    if not surrounding_prices or atr <= 0:
        return 0.0
    if high:
        distance = price - max(surrounding_prices)
    else:
        distance = min(surrounding_prices) - price
    return max(distance / atr, 0.0)


def _write_csv(path: Path, rows: list[dict[str, object]]) -> Path:
    if not rows:
        path.write_text("", encoding="utf-8")
        return path
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    return path


def _write_xlsx(path: Path, rows: list[dict[str, object]]) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except Exception:
        return _write_csv(path.with_suffix(".csv"), rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "swings"
    _append_rows(ws, rows)
    if rows:
        table = Table(displayName="Swings", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)

    ws_summary = wb.create_sheet("summary")
    _append_rows(ws_summary, swing_summary_rows(rows))
    for sheet in wb.worksheets:
        _format_sheet(sheet)
    wb.save(path)
    return path


def _append_rows(ws, rows: list[dict[str, object]]) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])


def _format_sheet(ws) -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
    ws.freeze_panes = "A2"
    for col in range(1, ws.max_column + 1):
        width = max(len(str(ws.cell(row, col).value or "")) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), 32)


def parse_timeframes(value: str) -> tuple[str, ...]:
    return tuple(item.strip().upper() for item in value.split(",") if item.strip())


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export confirmed swing high/low points from MT5 bridge history.")
    parser.add_argument("--history", default="runtime/latest_history_168h.json")
    parser.add_argument("--output", default="reports/swing_points.xlsx")
    parser.add_argument("--timeframes", default="M1,M5")
    parser.add_argument("--left", type=int, default=None, help="Override left-side swing window for all timeframes.")
    parser.add_argument("--right", type=int, default=None, help="Override right-side confirmation window for all timeframes.")
    parser.add_argument("--min-atr-distance", type=float, default=0.5)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    history = load_history(args.history)
    rows = swing_report_rows(
        history,
        timeframes=parse_timeframes(args.timeframes),
        left=args.left,
        right=args.right,
        min_atr_distance=args.min_atr_distance,
    )
    written = write_swing_report(args.output, rows)
    summary = swing_summary_rows(rows)
    print({"rows": len(rows), "summary": summary})
    print(f"wrote {written}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
