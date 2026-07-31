from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

if __package__ is None or __package__ == "":
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from analysis.market_data import Bar, MarketHistory, TIME_FORMAT, index_at_or_before, load_history


DEAL_TIME_FORMATS = ("%Y.%m.%d %H:%M:%S", TIME_FORMAT)


def load_deal_history(path: str | Path) -> dict[str, Any]:
    with Path(path).open(encoding="utf-8") as f:
        payload = json.load(f)
    return payload if isinstance(payload, dict) else {"deals": payload}


def deal_context_rows(
    history: MarketHistory,
    deal_history: dict[str, Any],
    *,
    symbol: str | None = None,
    entry_filter: str = "out",
    before_minutes: int = 10,
    after_minutes: int = 10,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    bars = history.bars("M1")
    if not bars:
        return [], []
    selected_deals = filter_deals(deal_history.get("deals") or [], symbol=symbol, entry_filter=entry_filter)
    summaries: list[dict[str, object]] = []
    contexts: list[dict[str, object]] = []
    for deal in selected_deals:
        deal_time = parse_deal_time(deal.get("time"))
        if deal_time is None:
            continue
        bar_index = index_at_or_before(bars, deal_time)
        if bar_index is None:
            continue
        start = max(0, bar_index - before_minutes)
        end = min(len(bars) - 1, bar_index + after_minutes)
        deal_summary = summarize_deal_context(deal, bars, bar_index, start, end)
        summaries.append(deal_summary)
        for index in range(start, end + 1):
            contexts.append(context_row(deal, bars[index], index - bar_index, index == bar_index))
    return summaries, contexts


def filter_deals(
    deals: list[dict[str, Any]],
    *,
    symbol: str | None,
    entry_filter: str,
) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for deal in deals:
        if symbol and str(deal.get("symbol") or "") != symbol:
            continue
        entry = str(deal.get("entry") or "").lower()
        if entry_filter != "all" and entry != entry_filter:
            continue
        if parse_deal_time(deal.get("time")) is None:
            continue
        result.append(deal)
    return sorted(result, key=lambda item: parse_deal_time(item.get("time")) or datetime.min)


def summarize_deal_context(
    deal: dict[str, Any],
    bars: list[Bar],
    bar_index: int,
    start: int,
    end: int,
) -> dict[str, object]:
    window = bars[start : end + 1]
    before = bars[start:bar_index]
    after = bars[bar_index + 1 : end + 1]
    price = number(deal.get("price"))
    profit = net_profit(deal)
    return {
        **deal_fields(deal),
        "context_start": bars[start].time_text,
        "context_end": bars[end].time_text,
        "matched_bar_time": bars[bar_index].time_text,
        "context_bars": len(window),
        "before_bars": len(before),
        "after_bars": len(after),
        "window_high": round(max(bar.high for bar in window), 5),
        "window_low": round(min(bar.low for bar in window), 5),
        "window_range": round(max(bar.high for bar in window) - min(bar.low for bar in window), 5),
        "pre_close_change": round(bars[bar_index].close - before[0].open, 5) if before else 0.0,
        "post_close_change": round(after[-1].close - bars[bar_index].close, 5) if after else 0.0,
        "close_from_deal_price": round(bars[bar_index].close - price, 5) if price is not None else "",
        "net_profit": profit,
    }


def context_row(deal: dict[str, Any], bar: Bar, relative_minute: int, matched: bool) -> dict[str, object]:
    price = number(deal.get("price"))
    return {
        **deal_fields(deal),
        "relative_minute": relative_minute,
        "matched_deal_minute": matched,
        "bar_time": bar.time_text,
        "bar_open": bar.open,
        "bar_high": bar.high,
        "bar_low": bar.low,
        "bar_close": bar.close,
        "bar_range": round(bar.high - bar.low, 5),
        "bar_body": round(bar.close - bar.open, 5),
        "tick_volume": bar.tick_volume,
        "close_from_deal_price": round(bar.close - price, 5) if price is not None else "",
        "high_from_deal_price": round(bar.high - price, 5) if price is not None else "",
        "low_from_deal_price": round(bar.low - price, 5) if price is not None else "",
    }


def deal_fields(deal: dict[str, Any]) -> dict[str, object]:
    return {
        "ticket": deal.get("ticket", ""),
        "deal_time": str(deal.get("time") or ""),
        "symbol": str(deal.get("symbol") or ""),
        "type": str(deal.get("type") or ""),
        "entry": str(deal.get("entry") or ""),
        "volume": number(deal.get("volume")) or 0.0,
        "deal_price": number(deal.get("price")) or 0.0,
        "profit": number(deal.get("profit")) or 0.0,
        "commission": number(deal.get("commission")) or 0.0,
        "swap": number(deal.get("swap")) or 0.0,
        "magic": deal.get("magic", ""),
    }


def parse_deal_time(value: object) -> datetime | None:
    if value is None:
        return None
    text = str(value)
    for fmt in DEAL_TIME_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def net_profit(deal: dict[str, Any]) -> float:
    return round((number(deal.get("profit")) or 0.0) + (number(deal.get("commission")) or 0.0) + (number(deal.get("swap")) or 0.0), 2)


def number(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def write_deal_context_report(
    path: str | Path,
    summaries: list[dict[str, object]],
    contexts: list[dict[str, object]],
) -> Path:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.suffix.lower() == ".xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except Exception:
            return write_csv(output.with_suffix(".csv"), contexts)

        wb = Workbook()
        for sheet_index, (title, rows) in enumerate((("summary", summaries), ("context", contexts))):
            ws = wb.active if sheet_index == 0 else wb.create_sheet(title)
            ws.title = title
            append_rows(ws, rows)
            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="305496")
                ws.freeze_panes = "A2"
            for col in range(1, ws.max_column + 1):
                width = max(len(str(ws.cell(row, col).value or "")) for row in range(1, ws.max_row + 1))
                ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), 36)
            if ws.max_row > 1 and ws.max_column > 1:
                table = Table(displayName=f"DealContext_{sheet_index}_{title}", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
                table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
                ws.add_table(table)
        wb.save(output)
        return output
    return write_csv(output, contexts)


def append_rows(ws, rows: list[dict[str, object]]) -> None:
    if not rows:
        return
    headers = ordered_headers(rows)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def write_csv(path: Path, rows: list[dict[str, object]]) -> Path:
    if not rows:
        path.write_text("", encoding="utf-8")
        return path
    headers = ordered_headers(rows)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
    return path


def ordered_headers(rows: list[dict[str, object]]) -> list[str]:
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    return headers


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Join MT5 closed deals with surrounding M1 bars.")
    parser.add_argument("--history", default="runtime/latest_history_168h.json")
    parser.add_argument("--deal-history", default="runtime/latest_deal_history.json")
    parser.add_argument("--symbol", default="")
    parser.add_argument("--entry", choices=("out", "in", "all"), default="out")
    parser.add_argument("--before-minutes", type=int, default=10)
    parser.add_argument("--after-minutes", type=int, default=10)
    parser.add_argument("--output", default="reports/deal_m1_context.xlsx")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    if args.before_minutes < 0 or args.after_minutes < 0:
        raise SystemExit("--before-minutes and --after-minutes must be >= 0")
    history = load_history(args.history)
    deal_history = load_deal_history(args.deal_history)
    summaries, contexts = deal_context_rows(
        history,
        deal_history,
        symbol=args.symbol or history.symbol or None,
        entry_filter=args.entry,
        before_minutes=args.before_minutes,
        after_minutes=args.after_minutes,
    )
    written = write_deal_context_report(args.output, summaries, contexts)
    print(f"deals: {len(summaries)}")
    print(f"context rows: {len(contexts)}")
    print(f"wrote {written}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
